"""One-sheet three-statement financial model export (annual + quarterly + LTM).

Layout follows the analyst's Financial_Model_Template.xlsx: line items down
column A, one column per fiscal year, then the **last four fiscal quarters**
(spanning the fiscal-year boundary, ``Q3'25 Q4'25 Q1'26 Q2'26`` style), then
**LTM** — with the income statement, balance sheet and cash-flow statement
consolidated on ONE sheet, styled in the house colour scheme.

The sheet adapts to each company's own presentation (as tagged in its SEC
filings): line items the filer never reports are dropped rather than left
blank, opex appears as the split lines (Sales & Marketing / G&A) or the
combined SG&A line — whichever the company files — and row labels follow
the winning XBRL tag (e.g. "Technology & Development").

Quarterly mechanics (as-filed XBRL under the annual winning tag):

- **discrete quarter** = the filed ~3-month duration when present; else
  fiscal-YTD differencing (10-Q cash-flow statements are YTD-only); a
  fiscal Q4 (never filed discretely) = FY − 9-month YTD (or FY − ΣQ1..Q3);
- **LTM (flows)** = last FY + latest fiscal YTD − year-ago comparative YTD
  (= the FY itself when the latest period end is the FY end);
- **balance-sheet rows** show period-end balances, latest in the LTM column;
- **per-share rows** use the same additive arithmetic (approximation).

**% change rows** sit under the relevant line items: fiscal-year cells are
YoY vs the prior fiscal year, and quarter cells are YoY vs the same fiscal
quarter a year earlier (seasonal businesses make quarter-on-quarter noise,
so QoQ is deliberately not shown). Computed from the same consolidated
values; year-ago quarters derive from the filing history exactly like the
displayed ones.

Everything is computed from data already fetched for the dashboard — the
export itself never touches the network.
"""
from __future__ import annotations

import datetime as dt
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

from . import config
from . import palette as P
from .edgar import (
    INSTANT_TAGS, AnnualFundamentals, QuarterlyFundamentals,
    parse_quarterly_facts,
)
from .market import summary_stat
from .metrics import DashboardData, fy_label
# FIX-15a: the quarter spine and discrete/YTD/LTM derivations moved verbatim
# to quarters.py (shared with the Explore TTM series); this suite is their
# behavior-freeze regression
from .quarters import (
    _SPAN_TOL, _discrete, _find_span, _ltm_flow, _match_instant,
    quarter_label, quarter_spine,
)
from .segments import partial_axis_disclosure

_MM = 1e6

# FIX-12h number formats by row kind (zero prints as a dash, negatives in
# parens for money/per-share; % rows carry an explicit sign) — defined once
# here at FIX-13d time, shared by the Model, as-filed and Segments sheets.
_FMT_MONEY = '#,##0.0;(#,##0.0);"–"'    # $mm rows
_FMT_PS = '0.00;(0.00);"–"'             # per-share rows (eps_diluted)
_FMT_SHARES = "#,##0.0"                 # share-count rows (never negative)
_FMT_PCT = '+0.0%;-0.0%;"–"'            # % change and tie-check rows
_FMT_RATIO = '0.0"×";-0.0"×";"–"'       # market multiples (FIX-16b)

# (label, concept, style, pct_row); concept None = section header;
# "=name" = derived row; style: item | total | eps | shares
LAYOUT: List[Tuple[str, Optional[str], str, bool]] = [
    ("INCOME STATEMENT ($mm; EPS in $, shares in mm)", None, "section", False),
    ("Total Revenue", "revenue", "item", True),
    ("Cost of Revenue", "cost_of_revenue", "item", True),
    ("Gross Profit", "=gross_profit", "total", True),
    ("Research & Development", "rnd", "item", True),
    ("Sales & Marketing", "marketing", "item", True),
    ("General & Administrative", "ga", "item", True),
    ("Selling, General & Administrative", "sga", "item", True),
    ("Total Operating Expenses", "opex_total", "item", True),
    ("Operating Income (EBIT)", "operating_income", "total", True),
    ("Interest Expense", "interest_expense", "item", False),
    ("Income Before Taxes", "=pretax", "item", False),
    ("Income Tax Provision", "tax_expense", "item", False),
    ("Net Income", "net_income", "total", True),
    ("Diluted EPS ($)", "eps_diluted", "eps", True),
    ("Diluted Shares (mm)", "diluted_shares", "shares", False),
    ("BALANCE SHEET (period end, $mm)", None, "section", False),
    ("Cash & Equivalents", "cash", "item", False),
    ("Accounts Receivable", "accounts_receivable", "item", False),
    ("Inventory", "inventory", "item", False),
    ("Total Current Assets", "assets_current", "total", False),
    ("Property & Equipment, Net", "ppe_net", "item", False),
    ("Goodwill", "goodwill", "item", False),
    ("Total Assets", "total_assets", "total", False),
    ("Accounts Payable", "accounts_payable", "item", False),
    ("Short-Term Borrowings", "st_borrowings", "item", False),
    ("Current Portion of Long-Term Debt", "lt_debt_current", "item", False),
    ("Total Current Liabilities", "liabilities_current", "total", False),
    ("Long-Term Debt", "lt_debt_noncurrent", "item", False),
    ("Total Liabilities", "liabilities_total", "total", False),
    ("Retained Earnings", "retained_earnings", "item", False),
    ("Minority Interest", "minority_interest", "item", False),
    ("Preferred Equity", "preferred_equity", "item", False),
    ("Total Stockholders' Equity", "equity", "total", False),
    ("CASH FLOW STATEMENT ($mm)", None, "section", False),
    ("Net Income", "net_income", "item", False),
    ("Depreciation & Amortization", "dna", "item", False),
    ("Stock-Based Compensation", "sbc", "item", False),
    ("Cash from Operations", "cfo", "total", True),
    ("Capital Expenditure", "capex", "item", True),
    ("Cash from Investing", "cfi", "total", False),
    ("Dividends Paid", "dividends_paid", "item", False),
    ("Share Repurchases", "buybacks", "item", False),
    ("Cash from Financing", "cff", "total", False),
    ("Free Cash Flow (CFO − capex)", "=fcf", "total", True),
]

# Row labels follow the company's presentation when the winning tag says so
_TAG_LABELS = {
    "TechnologyAndDevelopmentExpense": "Technology & Development",
    "AdvertisingExpense": "Advertising",
    "CostsAndExpenses": "Total Costs & Expenses",
    "CostOfGoodsAndServicesSold": "Cost of Goods & Services Sold",
}

# Weighted-average share counts can't be derived by FY − YTD subtraction
_NO_FY_DIFF = {"diluted_shares", "basic_shares"}


@dataclass
class ModelRow:
    """One consolidated line: annual, displayed quarters, LTM (+ internals)."""

    ann: List[Optional[float]]
    q: List[Optional[float]]                 # aligned to the displayed spine
    ltm: Optional[float]
    # same-quarter-a-year-earlier values, aligned to q (drives the YoY cells)
    q_prior: List[Optional[float]] = field(default_factory=list)
    # FIX-11c provenance: "ltm" (true trailing twelve months), "fy" (fell
    # back to the completed fiscal year), "mixed" (derived row whose legs
    # disagree — value suppressed), "none"
    ltm_basis: str = "none"
    ltm_date: Optional[dt.date] = None       # balance date (instant rows)


def _latest(vals: List[Optional[float]]) -> Optional[float]:
    for v in reversed(vals):
        if v is not None:
            return v
    return None


def _pct(cur: Optional[float], prev: Optional[float]) -> Optional[float]:
    if cur is None or prev is None or prev <= 0:
        return None
    return cur / prev - 1.0


def _annual_from_entries(entries, fy_ends: List[dt.date]
                         ) -> List[Optional[float]]:
    """Full-year values aligned to the fiscal spine (segment lines)."""
    return [next((v for s, e, v in entries
                  if abs((e - fe).days) <= 7 and 330 <= (e - s).days <= 400),
                 None)
            for fe in fy_ends]


def _annual_spans(entries, fy_ends: List[dt.date]):
    """The matched full-year span per fiscal year (for synth-flagging)."""
    return [next(((s, e) for s, e, v in entries
                  if abs((e - fe).days) <= 7 and 330 <= (e - s).days <= 400),
                 None)
            for fe in fy_ends]


def _flow_row(entries, ann: List[Optional[float]], fy_ends: List[dt.date],
              q_ends: List[dt.date], allow_fy_diff: bool = True,
              shares_like: bool = False) -> ModelRow:
    """Consolidate one flow line: quarters, year-ago quarters, LTM."""
    annual_map = dict(zip(fy_ends, ann))
    qs = [_discrete(entries, qe, fy_ends, annual_map,
                    allow_fy_diff=allow_fy_diff) for qe in q_ends]
    # same fiscal quarter a year earlier — the filing history (prior-year
    # 10-Qs and comparatives) supports the same derivation chain
    q_prior = [_discrete(entries, qe - dt.timedelta(days=365), fy_ends,
                         annual_map, allow_fy_diff=allow_fy_diff)
               for qe in q_ends]
    if shares_like:
        ltm = _latest(qs)  # latest weighted count
        basis = "ltm" if ltm is not None else "none"
        if ltm is None:
            ltm = _latest(ann)
            basis = "fy" if ltm is not None else "none"
    else:
        ltm, basis = _ltm_flow(ann[-1] if ann else None, entries,
                               fy_ends, q_ends)
    return ModelRow(ann, qs, ltm, q_prior, ltm_basis=basis)


# ----------------------------------------------------------- consolidation

def build_model_rows(annual: AnnualFundamentals,
                     qdata: QuarterlyFundamentals) -> Tuple[
                         Dict[str, ModelRow], List[dt.date], List[dt.date]]:
    """concept -> ModelRow, plus the annual and displayed-quarter spines."""
    fy_ends = annual.fy_ends
    q_ends = quarter_spine(qdata, fy_ends)

    rows: Dict[str, ModelRow] = {}
    concepts = {c for _, c, _, _ in LAYOUT if c and not c.startswith("=")}
    concepts.update(("gross_profit", "pretax_income"))  # feed derived rows
    for concept in concepts:
        ann = list(annual.series.get(concept) or [None] * len(fy_ends))
        if concept in INSTANT_TAGS:
            obs = qdata.instant.get(concept, {})
            qs = [_match_instant(obs, qe) for qe in q_ends]
            ltm = _latest(qs)
            ltm_date = next((qe for qe, v in zip(reversed(q_ends),
                                                 reversed(qs))
                             if v is not None), None)
            if ltm is None:
                ltm = _latest(ann)  # latest period-end balance
                ltm_date = next((fe for fe, v in zip(reversed(fy_ends),
                                                     reversed(ann))
                                 if v is not None), None)
            rows[concept] = ModelRow(ann, qs, ltm, [None] * len(q_ends),
                                     ltm_basis="ltm" if ltm is not None
                                     else "none", ltm_date=ltm_date)
            continue
        rows[concept] = _flow_row(
            qdata.duration.get(concept, []), ann, fy_ends, q_ends,
            allow_fy_diff=concept not in _NO_FY_DIFF,
            shares_like=concept in _NO_FY_DIFF)

    def _combine(a: ModelRow, b: ModelRow, op) -> ModelRow:
        def cell(x, y):
            return op(x, y) if x is not None and y is not None else None
        # FIX-11c: an LTM only combines when both legs share a basis — a
        # trailing-twelve-month leg minus a fiscal-year leg is a number
        # that means nothing; suppress it and say so (basis "mixed")
        if a.ltm is None or b.ltm is None:
            ltm, basis = None, "none"
        elif a.ltm_basis == b.ltm_basis:
            ltm, basis = op(a.ltm, b.ltm), a.ltm_basis
        else:
            ltm, basis = None, "mixed"
        # prior-of-a-difference == difference-of-priors, so q_prior combines
        # element-wise exactly like the displayed quarters
        return ModelRow([cell(x, y) for x, y in zip(a.ann, b.ann)],
                        [cell(x, y) for x, y in zip(a.q, b.q)],
                        ltm,
                        [cell(x, y) for x, y in zip(a.q_prior, b.q_prior)],
                        ltm_basis=basis)

    def _prefer(tagged: ModelRow, derived: ModelRow) -> ModelRow:
        def cell(t, d):
            return t if t is not None else d
        return ModelRow([cell(t, d) for t, d in zip(tagged.ann, derived.ann)],
                        [cell(t, d) for t, d in zip(tagged.q, derived.q)],
                        cell(tagged.ltm, derived.ltm),
                        [cell(t, d) for t, d in
                         zip(tagged.q_prior, derived.q_prior)],
                        ltm_basis=(tagged.ltm_basis if tagged.ltm is not None
                                   else derived.ltm_basis))

    sub = lambda x, y: x - y  # noqa: E731
    add = lambda x, y: x + y  # noqa: E731
    # Gross profit: as tagged, falling back to revenue − cost of revenue
    rows["=gross_profit"] = _prefer(
        rows["gross_profit"], _combine(rows["revenue"],
                                       rows["cost_of_revenue"], sub))
    # Pre-tax income: as tagged, falling back to net income + tax provision
    rows["=pretax"] = _prefer(
        rows["pretax_income"], _combine(rows["net_income"],
                                        rows["tax_expense"], add))
    rows["=fcf"] = _combine(rows["cfo"], rows["capex"], sub)  # capex = +outflow
    return rows, fy_ends, q_ends


# ------------------------------------------------------------------ writer

def _write_note_row(ws, r: int, text: str) -> int:
    """Single muted note row (FIX-14d: stands in for a suppressed tie)."""
    from openpyxl.styles import Font
    muted = P.INK_MUTED.lstrip("#").upper()
    ws.cell(row=r, column=1, value=text)
    ws.cell(row=r, column=1).font = Font(italic=True, color=muted, size=8.5)
    return r + 1


def _tie_note(n_members: int, sigma, total) -> str:
    cvg = (f"{sigma / total:.0%}" if sigma is not None and total
           else "unknown share")
    return (f"   partial disclosure axis — tie suppressed "
            f"({n_members} member(s), {cvg} of consolidated)")


def _write_check_row(ws, r: int, label: str, values, ltm_col: int,
                     pct_fmt: bool = False, tol: float = 0.02) -> int:
    """Muted checksum row (segment Σ/gap, income-statement tie): values
    beyond `tol` render in the flag red. Shared by the IS tie row and the
    segments block."""
    from openpyxl.styles import Alignment, Font
    muted = P.INK_MUTED.lstrip("#").upper()
    bad = P.DELTA_BAD.lstrip("#").upper()
    ws.cell(row=r, column=1, value=label)
    ws.cell(row=r, column=1).font = Font(italic=True, color=muted, size=8.5)
    for j, v in enumerate(values, start=2):
        c = ws.cell(row=r, column=j)
        color = muted
        if v is not None:
            c.value = v if pct_fmt else v / _MM
            c.number_format = _FMT_PCT if pct_fmt else _FMT_MONEY  # FIX-12h
            if pct_fmt and abs(v) > tol:
                color = bad  # the tie is off — make it read as a flag
        c.font = Font(italic=True, color=color, size=8.5)
        c.alignment = Alignment(horizontal="right")
    return r + 1


def _label_for(default: str, concept: Optional[str],
               annual: AnnualFundamentals) -> str:
    if not concept or concept.startswith("="):
        return default
    primary = (annual.tags_used.get(concept) or "").split(" (")[0]
    return _TAG_LABELS.get(primary, default)


def export_financial_model(d: DashboardData, path: str) -> str:
    """Write the consolidated three-statement model workbook; returns path."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    f = getattr(d, "fundamentals", None)
    if f is None or f.raw_facts is None:
        raise ValueError(
            "financial-model export needs the fetched filing data — "
            "run Analyze on a ticker first")
    qdata = parse_quarterly_facts(f.raw_facts, f)
    rows, fy_ends, q_ends = build_model_rows(f, qdata)

    headers = ([fy_label(e) for e in fy_ends]
               + [quarter_label(qe, fy_ends) for qe in q_ends]
               + ["LTM"])
    ltm_col = 1 + len(headers)  # 1-based; column A is the line items
    headers.append("CAGR/avg")  # FIX-16b per-row summary column
    sum_col = ltm_col + 1

    ink = P.INK_PRIMARY.lstrip("#").upper()
    forest = P.GUI_SIDEBAR_BG.lstrip("#").upper()
    cream = P.SURFACE.lstrip("#").upper()
    muted = P.INK_MUTED.lstrip("#").upper()
    section_fill = PatternFill("solid", fgColor="DFE9E1")   # pale sage
    ltm_fill = PatternFill("solid", fgColor="FBF0D4")       # pale amber
    header_fill = PatternFill("solid", fgColor=forest)
    total_border = Border(top=Side(style="thin", color="9AA79B"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Model"
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False

    ws.cell(row=1, column=1, value="Line Items")
    for j, h in enumerate(headers, start=2):
        ws.cell(row=1, column=j, value=h)
    for j in range(1, sum_col + 1):
        c = ws.cell(row=1, column=j)
        c.fill = header_fill
        c.font = Font(bold=True, color=cream, size=10)
        c.alignment = Alignment(horizontal="left" if j == 1 else "right")

    # FIX-12h: formats chosen per row kind; the segments block below reuses
    # the same money/percent formats through these locals
    fmt_mm, fmt_eps, fmt_pct = _FMT_MONEY, _FMT_PS, _FMT_PCT
    is_tie_breached = False  # set by the FIX-11b Rev−COGS-vs-GP tie row
    rendered_rows: List[Tuple[str, str, str]] = []  # (label, concept, style)

    def pct_cells(row: ModelRow) -> List[Optional[float]]:
        """FY cells: YoY vs prior FY · quarter cells: YoY vs the same
        fiscal quarter a year earlier · LTM cell: blank."""
        fy = [None] + [_pct(row.ann[i], row.ann[i - 1])
                       for i in range(1, len(row.ann))]
        qs = [_pct(v, p) for v, p in zip(row.q, row.q_prior)]
        return fy + qs + [None]

    def write_pct_row(r: int, row: ModelRow) -> int:
        pcts = pct_cells(row)
        if all(v is None for v in pcts):
            return r
        ws.cell(row=r, column=1, value="   % change")
        ws.cell(row=r, column=1).font = Font(italic=True, color=muted,
                                             size=8.5)
        for j, v in enumerate(pcts, start=2):
            c = ws.cell(row=r, column=j)
            if v is not None:
                c.value = v
                c.number_format = fmt_pct
            c.font = Font(italic=True, color=muted, size=8.5)
            c.alignment = Alignment(horizontal="right")
        return r + 1

    r = 2
    for label, concept, style, pct_row in LAYOUT:
        if style == "section":
            ws.cell(row=r, column=1, value=label)
            for j in range(1, sum_col + 1):
                c = ws.cell(row=r, column=j)
                c.fill = section_fill
                c.font = Font(bold=True, color=forest, size=10)
            r += 1
            continue
        row = rows[concept]
        cells = list(row.ann) + list(row.q) + [row.ltm]
        if all(v is None for v in cells):
            continue  # the filer never reports this line — drop the row
        disp = _label_for(label, concept, f)
        rendered_rows.append((disp, concept, style))
        ws.cell(row=r, column=1, value=disp)
        scale = 1.0 if style == "eps" else _MM
        numfmt = {"eps": fmt_eps, "shares": _FMT_SHARES}.get(style, fmt_mm)
        bold = style == "total"
        ws.cell(row=r, column=1).font = Font(bold=bold, color=ink, size=10)
        for j, v in enumerate(cells, start=2):
            c = ws.cell(row=r, column=j)
            if v is not None:
                c.value = v / scale
                c.number_format = numfmt
            c.font = Font(bold=bold, color=ink, size=10)
            c.alignment = Alignment(horizontal="right")
            if j == ltm_col:
                c.fill = ltm_fill
            if bold:
                c.border = total_border
        if bold:
            ws.cell(row=r, column=1).border = total_border
        # FIX-16b: per-row summary — geometric CAGR over the annual cells
        # (None on sign-flips or thin data, matching the anchor-ladder rule)
        s = summary_stat(list(row.ann), "cagr")
        if s is not None:
            c = ws.cell(row=r, column=sum_col, value=s)
            c.number_format = fmt_pct
            c.font = Font(italic=True, color=muted, size=8.5)
            c.alignment = Alignment(horizontal="right")
        r += 1
        if pct_row:
            r = write_pct_row(r, row)
        if concept == "=gross_profit":
            # FIX-11b: the income statement referees itself on the sheet's
            # face — Revenue − Cost of Revenue must equal the TAGGED Gross
            # Profit; a breach means the three lines mix accounting bases
            # (MELI: headline vs contract-only revenue).
            tag_gp, rev_r, cogs_r = (rows["gross_profit"], rows["revenue"],
                                     rows["cost_of_revenue"])

            def _cells(mr: ModelRow):
                return list(mr.ann) + list(mr.q) + [mr.ltm]

            if all(any(v is not None for v in _cells(mr))
                   for mr in (tag_gp, rev_r, cogs_r)):
                gaps = [((rv - cg - gp_) / rv
                         if None not in (rv, cg, gp_) and rv else None)
                        for rv, cg, gp_ in zip(_cells(rev_r), _cells(cogs_r),
                                               _cells(tag_gp))]
                if any(g is not None for g in gaps):
                    r = _write_check_row(ws, r, "   Rev − COGS vs GP (gap)",
                                         gaps, ltm_col, pct_fmt=True,
                                         tol=config.IS_TIE_TOL)
                    is_tie_breached = any(
                        g is not None and abs(g) > config.IS_TIE_TOL
                        for g in gaps)


    # --------------------------------------------- MARKET & RATIOS (FIX-16b)
    # Fundamentals joined to the FY-end market series, per year — the
    # value investor's one-glance table (owner-ratified benchmark insight).
    # The LTM column holds TODAY's values; years before the price window
    # stay blank. Ratios mask on non-positive denominators (market.py).
    disp_by_end = {e: i for i, e in enumerate(d.fy_ends)}

    def _mk_full(vals_disp):
        return [vals_disp[disp_by_end[fe]]
                if fe in disp_by_end and disp_by_end[fe] < len(vals_disp)
                else None for fe in fy_ends]

    shares_now = _latest(d.diluted_shares)
    mcap_now = (d.last_close * shares_now
                if d.last_close and shares_now else None)
    nd_now = _latest(d.net_debt_fy)
    mi_now = _latest(d.minority_interest) or 0.0
    pref_now = _latest(d.preferred_equity) or 0.0
    ev_now = (mcap_now + nd_now + mi_now + pref_now
              if mcap_now is not None and nd_now is not None else None)
    eps_now = _latest(d.eps_diluted)
    ebit_now = _latest(d.ebit_reported)
    market_rows = [
        ("Market Cap", d.market_cap_fy, "money", mcap_now, "cagr"),
        ("Enterprise Value", d.ev_fy, "money", ev_now, "cagr"),
        ("Net Debt", d.net_debt_fy, "money", nd_now, None),
        ("Tangible Book (equity − goodwill − intangibles)",
         d.tangible_book, "money", _latest(d.tangible_book), "cagr"),
        ("P/E (FY-end close)", d.pe_fy, "ratio",
         (d.last_close / eps_now
          if d.last_close and eps_now and eps_now > 0 else None), "avg"),
        ("EV/EBIT", d.ev_ebit_fy, "ratio",
         (ev_now / ebit_now
          if ev_now is not None and ebit_now and ebit_now > 0 else None),
         "avg"),
        ("Net Debt/EBIT", d.net_debt_ebit_fy, "ratio",
         (nd_now / ebit_now
          if nd_now is not None and ebit_now and ebit_now > 0 else None),
         "avg"),
        ("Adj FCF Yield (ex-SBC / mkt cap)", d.adj_fcf_yield_fy, "pct",
         d.adj_fcf_yield_now, "avg"),
    ]
    market_written = False
    for mlabel, disp_vals, kind, now_val, sum_kind in market_rows:
        full = _mk_full(list(disp_vals or []))
        if all(v is None for v in full) and now_val is None:
            continue
        if not market_written:
            ws.cell(row=r, column=1,
                    value="MARKET & RATIOS (FY-end close × diluted shares; "
                          "EV incl. bridge legs; LTM column = today)")
            for j in range(1, sum_col + 1):
                c = ws.cell(row=r, column=j)
                c.fill = section_fill
                c.font = Font(bold=True, color=forest, size=10)
            r += 1
            market_written = True
        ws.cell(row=r, column=1, value=mlabel)
        ws.cell(row=r, column=1).font = Font(color=ink, size=10)
        mscale = _MM if kind == "money" else 1.0
        mfmt = {"money": fmt_mm, "ratio": _FMT_RATIO, "pct": fmt_pct}[kind]
        cells16 = full + [None] * len(q_ends) + [now_val]
        for j, v in enumerate(cells16, start=2):
            c = ws.cell(row=r, column=j)
            if v is not None:
                c.value = v / mscale
                c.number_format = mfmt
            c.font = Font(color=ink, size=10)
            c.alignment = Alignment(horizontal="right")
            if j == ltm_col:
                c.fill = ltm_fill
        if sum_kind:
            s16 = summary_stat(full, sum_kind)
            if s16 is not None:
                c = ws.cell(row=r, column=sum_col, value=s16)
                c.number_format = fmt_pct if sum_kind == "cagr" else mfmt
                c.font = Font(italic=True, color=muted, size=8.5)
                c.alignment = Alignment(horizontal="right")
        r += 1

    # ------------------------------------------------- SEGMENTS (as filed)
    seg = getattr(d, "segments", None)
    seg_lines = list(getattr(seg, "lines", None) or [])
    any_synth_cell = False
    if seg_lines:
        ws.cell(row=r, column=1, value="SEGMENTS (as filed)")
        for j in range(1, ltm_col + 1):
            c = ws.cell(row=r, column=j)
            c.fill = section_fill
            c.font = Font(bold=True, color=forest, size=10)
        r += 1

        blocks: List[Tuple[Tuple[str, str], List]] = []
        for ln in seg_lines:  # group into (measure, axis) blocks, in order
            key = (ln.group, ln.axis)
            if not blocks or blocks[-1][0] != key:
                blocks.append((key, []))
            blocks[-1][1].append(ln)
        for (group, axis), lns in blocks:
            rendered = []
            for ln in lns:
                ann = _annual_from_entries(ln.entries, fy_ends)
                rowv = _flow_row(ln.entries, ann, fy_ends, q_ends)
                if any(v is not None for v in
                       list(rowv.ann) + list(rowv.q) + [rowv.ltm]):
                    rendered.append((ln, rowv))
            if not rendered:
                continue
            c = ws.cell(row=r, column=1, value=f"{group} by {axis}")
            c.font = Font(bold=True, italic=True, color=forest, size=9)
            r += 1
            for ln, rowv in rendered:
                spans = _annual_spans(ln.entries, fy_ends)
                flags = ([sp is not None and sp in ln.synth for sp in spans]
                         + [any(abs((e - qe).days) <= _SPAN_TOL
                                for _, e in ln.synth) for qe in q_ends]
                         + [bool(ln.synth)
                            and any(e >= fy_ends[-1] for _, e in ln.synth)])
                cells = list(rowv.ann) + list(rowv.q) + [rowv.ltm]
                ws.cell(row=r, column=1, value=f"  {ln.member}")
                ws.cell(row=r, column=1).font = Font(color=ink, size=10)
                for j, (v, fl) in enumerate(zip(cells, flags), start=2):
                    c = ws.cell(row=r, column=j)
                    if v is not None:
                        c.value = v / _MM
                        c.number_format = fmt_mm
                        if fl:
                            any_synth_cell = True
                    c.font = Font(color=ink, size=10, italic=fl)
                    c.alignment = Alignment(horizontal="right")
                    if j == ltm_col:
                        c.fill = ltm_fill
                r += 1
                r = write_pct_row(r, rowv)
            if group != "Revenue":
                continue
            # visible tie-out (house scale-tie doctrine): Σ of the members
            # vs the consolidated statement, per column — a positive gap
            # means hierarchical members double-count on this axis
            cons = rows["revenue"]
            cons_cells = list(cons.ann) + list(cons.q) + [cons.ltm]
            sums: List[Optional[float]] = []
            for i in range(len(cons_cells)):
                vals = [(list(rv.ann) + list(rv.q) + [rv.ltm])[i]
                        for _, rv in rendered]
                vals = [v for v in vals if v is not None]
                sums.append(sum(vals) if vals else None)
            # FIX-14d: gate the tie on the latest column where members have
            # values (annual preferred) — a deliberately partial axis (one
            # member, sliver of revenue) would only wolf-cry a huge red gap
            with_data = [i for i, sv in enumerate(sums) if sv is not None]
            annual = [i for i in with_data if i < len(cons.ann)]
            gate_i = max(annual) if annual else (
                max(with_data) if with_data else None)
            n_mem = 0 if gate_i is None else sum(
                1 for _, rv in rendered
                if (list(rv.ann) + list(rv.q) + [rv.ltm])[gate_i] is not None)
            g_sigma = sums[gate_i] if gate_i is not None else None
            g_total = cons_cells[gate_i] if gate_i is not None else None
            if partial_axis_disclosure(n_mem, g_sigma, g_total):
                r = _write_note_row(ws, r, _tie_note(n_mem, g_sigma, g_total))
                continue
            r = _write_check_row(ws, r, "   Σ members", sums, ltm_col)
            r = _write_check_row(ws, r, "   vs consolidated (gap %)",
                                 [_pct(sv, cv) for sv, cv
                                  in zip(sums, cons_cells)], ltm_col,
                                 pct_fmt=True,
                                 tol=config.SEGMENT_TIE_TOL)  # FIX-10d

    # ---------------------------------------------------------- footnotes
    notes = [
        f"{d.company} ({d.ticker}) — consolidated financial model · "
        f"generated {d.generated.isoformat()} · USD in millions "
        "(EPS in $, shares in mm) · values as filed (SEC EDGAR XBRL; "
        "latest amendment wins).",
        "Quarter columns are the last four fiscal quarters: discrete "
        "3-month values as filed, else fiscal-YTD differencing (10-Q "
        "cash-flow statements are YTD-only); a fiscal Q4 is derived as "
        "FY − 9-month YTD (or FY − ΣQ1..Q3) — note a restated FY places "
        "the full restatement delta in that derived Q4.",
        "LTM (flows) = last FY + latest fiscal YTD − year-ago comparative "
        "YTD (= the FY itself when the latest period end is the FY end, or "
        "when a concept has no current-year interim data). Balance-sheet "
        "rows show the latest period-end balance in the LTM column; "
        "per-share rows use the same additive arithmetic (approximation).",
        "% change rows: fiscal-year cells are YoY vs the prior fiscal "
        "year; quarter cells are YoY vs the SAME fiscal quarter a year "
        "earlier (QoQ is not shown — seasonality makes it noise). Blank "
        "where the prior-period base is missing or non-positive.",
        "Line items the filer never tags are omitted, so the sheet mirrors "
        "the company's own SEC presentation. Derived rows: Gross Profit "
        "falls back to Revenue − Cost of Revenue; Income Before Taxes to "
        "Net Income + Tax Provision; Free Cash Flow = CFO − capex. Capex "
        "and other 'Payments…' concepts are positive outflows as filed.",
        "Not investment advice.",
    ]
    if market_written:  # FIX-16b provenance for the market block
        notes.insert(-1, (
            "MARKET & RATIOS: market cap = FY-end close × diluted shares "
            "(the same MVE read as Altman Z); EV adds net debt + minority "
            "interest + preferred (the equity-bridge legs); adj FCF = FCF "
            "ex-SBC (house §2b); the LTM column holds today's values at "
            "the last close."
            + (f" Owner's yield now (dividends + gross buybacks / market "
               f"cap; issuance NOT netted — see the dilution panel): "
               f"{d.owners_yield:.1%}." if d.owners_yield is not None
               else "")
            + " Years before the price window are blank; ratios mask on "
              "non-positive denominators. CAGR/avg column: geometric CAGR "
              "for value rows, period average for multiples and yields."))
    if seg_lines:
        src = getattr(seg, "source", "") or "latest filings"
        extra = getattr(seg, "status", "")
        notes.insert(-1, (
            f"Segments: dimensional XBRL parsed from {src}; history depth "
            "= as reported there (a 10-K carries 2–3 comparative years, "
            "the 10-Q the current quarters + year-ago comparatives). "
            "Members ordered by latest annual revenue. The 'Σ members' / "
            "'vs consolidated (gap %)' rows tie each revenue block to the "
            "consolidated statement — a positive gap beyond +2% signals "
            "hierarchical (parent + child) members double-counting on that "
            "axis; a negative gap means members the filer left untagged."
            + (f" Note: {extra}." if extra else "")))
        if any_synth_cell:
            notes.insert(-1, (
                "Italic segment cells were synthesized by summing the "
                "two-axis disaggregation table (not directly filed); the "
                "gap row shows how completely that table covers the "
                "consolidated total."))
        # FIX-10d: the audit trail rides in the footnotes
        cov = list(getattr(seg, "coverage", None) or [])
        if cov:
            matched = sum(1 for _, n in cov if n > 0)
            notes.insert(-1, (
                f"Segment coverage: dimensional facts found in {matched}/"
                f"{len(cov)} instances ({cov[0][0]} … {cov[-1][0]})."))
        for b in getattr(seg, "breaks", None) or []:
            notes.insert(-1, ("Segment recast — series are not comparable "
                              "across this boundary: " + b))
        recasts = list(getattr(seg, "recast_log", None) or [])
        if recasts:
            notes.insert(-1, (
                f"{len(recasts)} restated segment value(s) across filings: "
                + " | ".join(recasts[:3])
                + (" | …" if len(recasts) > 3 else "")))
        disc = [f"{ln.member} ({ln.group} by {ln.axis})" for ln in seg_lines
                if getattr(ln, "discontinuous", False)]
        if disc:
            notes.insert(-1, (
                "Discontinuous segment series (interior fiscal year missing "
                "while axis peers report one): " + ", ".join(disc) + "."))
    else:
        why = getattr(seg, "status", "") if seg is not None else \
            "not fetched in this run"
        notes.insert(-1, (
            "Segments: no dimensional segment data in this workbook — "
            + (why or "it lives in the 10-K/10-Q XBRL instance (fetched "
                      "live, not in the companyfacts API)") + "."))
    if is_tie_breached:
        notes.insert(-1, (
            "Income-statement basis check: Revenue − Cost of Revenue "
            "differs from Gross Profit beyond "
            f"±{config.IS_TIE_TOL:.0%} in the flagged columns — the three "
            "lines are not on one accounting basis; see the XBRL tags "
            "footnote."))
    # FIX-11c: LTM basis provenance on the sheet's face
    for lbl, c, _st in rendered_rows:
        if rows[c].ltm_basis == "mixed":
            notes.insert(-1, (
                f"LTM suppressed (mixed basis): {lbl} — one leg is "
                "trailing-twelve-month, the other has no current-year "
                "interim data."))
    if q_ends and q_ends[-1] > fy_ends[-1]:
        fy_rows = [lbl for lbl, c, st in rendered_rows
                   if c not in INSTANT_TAGS and st != "shares"
                   and rows[c].ltm_basis == "fy"]
        if fy_rows:
            notes.insert(-1, (
                f"LTM equals {fy_label(fy_ends[-1])} (no current-year "
                "interim data under any candidate tag): "
                + ", ".join(fy_rows) + "."))
        stale = [(lbl, rows[c].ltm_date) for lbl, c, _st in rendered_rows
                 if c in INSTANT_TAGS and rows[c].ltm_date is not None
                 and (q_ends[-1] - rows[c].ltm_date).days > _SPAN_TOL]
        if stale:
            notes.insert(-1, (
                "Balance shown is the latest tagged period-end, older than "
                "the newest quarter column: "
                + ", ".join(f"{lbl} ({d.isoformat()})" for lbl, d in stale)
                + "."))
    for note in qdata.source_notes:
        notes.insert(-1, "Interim gap-fill: " + note)
    tagged_pt, merged_pt = rows["pretax_income"], rows["=pretax"]
    if any(t is None and m is not None for t, m in zip(
            list(tagged_pt.ann) + list(tagged_pt.q) + [tagged_pt.ltm],
            list(merged_pt.ann) + list(merged_pt.q) + [merged_pt.ltm])):
        notes.insert(-1, (
            "Income Before Taxes cells without a filed pretax tag are "
            "derived as Net Income + Tax Provision; parent-attributable "
            "net income understates pretax income when minority interest "
            "is material."))
    if not getattr(d, "statements", None):  # FIX-13d degradation path
        why = getattr(d, "statements_note", "") or "not fetched in this run"
        notes.insert(-1, f"As-filed statement sheets unavailable: {why}")
    tag_bits = "; ".join(f"{k} = {v}" for k, v in sorted(f.tags_used.items()))
    if tag_bits:
        notes.append(f"XBRL tags: {tag_bits}")
    r += 1
    for note in notes:
        c = ws.cell(row=r, column=1, value=note)
        c.font = Font(color=muted, size=8)
        r += 1

    ws.column_dimensions["A"].width = 36
    for j in range(2, ltm_col + 1):
        ws.column_dimensions[get_column_letter(j)].width = 11.5

    # FIX-13d: the workbook becomes the staging layer — every tagged line
    # of the three primary statements in the filer's own order, plus the
    # per-axis Segments sheet; downstream extraction reads from here.
    _write_statement_sheets(wb, d)
    _write_segments_sheet(wb, d, qdata)
    wb.save(path)
    return path


def _unit_format(concept: str, unit: str) -> Tuple[float, str]:
    """(scale divisor, number format) per row kind — unit-aware."""
    if unit == "USD/shares" or concept.startswith("EarningsPerShare"):
        return 1.0, _FMT_PS
    if unit == "shares":
        return _MM, _FMT_SHARES
    if unit == "pure":
        return 1.0, "0.0000"
    return _MM, _FMT_MONEY


def _span_label(start: dt.date, end: dt.date) -> str:
    """Column label for an as-reported segment span."""
    days = (end - start).days
    if 330 <= days <= 400:
        return fy_label(end)
    return f"{round(days / 91) * 3}m to {end.isoformat()}"


def _write_statement_sheets(wb, d: DashboardData) -> None:
    """FIX-13d: 'Income Statement' / 'Balance Sheet' / 'Cash Flow' sheets —
    one row per presentation-linkbase line in the filer's own order and
    labels; values from companyfacts (latest amendment wins), raw as-filed
    signs (no negation games)."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from .edgar import annual_values_for_concept

    stmts = getattr(d, "statements", None)
    f = getattr(d, "fundamentals", None)
    if not stmts or f is None or f.raw_facts is None or not f.fy_ends:
        return
    short_names = stmts.get("_short_names", {})
    fy_ends = f.fy_ends
    ink = P.INK_PRIMARY.lstrip("#").upper()
    forest = P.GUI_SIDEBAR_BG.lstrip("#").upper()
    cream = P.SURFACE.lstrip("#").upper()
    muted = P.INK_MUTED.lstrip("#").upper()
    section_fill = PatternFill("solid", fgColor="DFE9E1")
    header_fill = PatternFill("solid", fgColor=forest)
    total_border = Border(top=Side(style="thin", color="9AA79B"))

    for key, title in (("income", "Income Statement"),
                       ("balance", "Balance Sheet"),
                       ("cashflow", "Cash Flow")):
        rows = stmts.get(key)
        if not rows:
            continue
        ws = wb.create_sheet(title)
        ws.freeze_panes = "B2"
        ws.sheet_view.showGridLines = False
        ws.cell(row=1, column=1, value="Line Items (as filed)")
        for j, e in enumerate(fy_ends, start=2):
            ws.cell(row=1, column=j, value=fy_label(e))
        for j in range(1, len(fy_ends) + 2):
            c = ws.cell(row=1, column=j)
            c.fill = header_fill
            c.font = Font(bold=True, color=cream, size=10)
            c.alignment = Alignment(horizontal="left" if j == 1 else "right")

        r = 2
        for row in rows:
            c1 = ws.cell(row=r, column=1, value="  " * row.depth + row.label)
            if row.is_abstract:  # bold section header, no values
                c1.font = Font(bold=True, color=forest, size=10)
                if row.depth == 0:
                    for j in range(1, len(fy_ends) + 2):
                        ws.cell(row=r, column=j).fill = section_fill
                r += 1
                continue
            vals, unit = annual_values_for_concept(
                f.raw_facts, row.concept, fy_ends)
            scale, numfmt = _unit_format(row.concept, unit)
            c1.font = Font(bold=row.is_total, color=ink, size=10)
            if row.is_total:
                c1.border = total_border
            for j, v in enumerate(vals, start=2):
                c = ws.cell(row=r, column=j)
                if v is not None:
                    c.value = v / scale
                    c.number_format = numfmt
                c.font = Font(bold=row.is_total, color=ink, size=10)
                c.alignment = Alignment(horizontal="right")
                if row.is_total:
                    c.border = total_border
            r += 1

        foot = [
            f"Source: 10-K {f.latest_10k_accession} · report "
            f"'{short_names.get(key, title)}'.",
            "Line order and labels as filed (presentation linkbase); values "
            "from SEC companyfacts, latest amendment wins; lines not in the "
            "current presentation are not shown.",
            "Raw as-filed signs: presentation-negated rows display the raw "
            "XBRL value (no sign flips applied). USD in millions; EPS in $; "
            "share counts in mm.",
        ]
        if key == "income":
            foot.append(
                "Operating KPIs disclosed in MD&A (GMV, TPV, NIMAL, active "
                "users/buyers, items sold, payment transactions) are not "
                "XBRL-tagged and are outside this export.")
        r += 1
        for note in foot:
            ws.cell(row=r, column=1, value=note).font = Font(color=muted,
                                                             size=8)
            r += 1
        ws.column_dimensions["A"].width = 52
        for j in range(2, len(fy_ends) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 11.5


def _write_segments_sheet(wb, d: DashboardData,
                          qdata: QuarterlyFundamentals) -> None:
    """FIX-13d: 'Segments' sheet — one block per axis (post-13b order),
    rows = member × measure, columns = the fiscal spans actually reported,
    synthesized cells italic, a Σ/tie pair per Revenue block vs the
    consolidated statement, and the full segment status as the footnote."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    seg = getattr(d, "segments", None)
    lines = list(getattr(seg, "lines", None) or [])
    if not lines:
        return
    ink = P.INK_PRIMARY.lstrip("#").upper()
    forest = P.GUI_SIDEBAR_BG.lstrip("#").upper()
    cream = P.SURFACE.lstrip("#").upper()
    muted = P.INK_MUTED.lstrip("#").upper()
    header_fill = PatternFill("solid", fgColor=forest)
    block_fill = PatternFill("solid", fgColor="DFE9E1")

    spans = sorted({(s, e) for ln in lines for s, e, _ in ln.entries},
                   key=lambda t: (t[1], t[0]))
    ws = wb.create_sheet("Segments")
    ws.freeze_panes = "B2"
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value="Member (as filed)")
    for j, (s, e) in enumerate(spans, start=2):
        ws.cell(row=1, column=j, value=_span_label(s, e))
    for j in range(1, len(spans) + 2):
        c = ws.cell(row=1, column=j)
        c.fill = header_fill
        c.font = Font(bold=True, color=cream, size=10)
        c.alignment = Alignment(horizontal="left" if j == 1 else "right")

    cons = qdata.duration.get("revenue", [])

    r = 2
    blocks: List[Tuple[Tuple[str, str], List]] = []
    for ln in lines:  # (measure, axis) blocks in post-13b line order
        key = (ln.group, ln.axis)
        if not blocks or blocks[-1][0] != key:
            blocks.append((key, []))
        blocks[-1][1].append(ln)
    for (group, axis), lns in blocks:
        c = ws.cell(row=r, column=1, value=f"{group} by {axis}")
        c.font = Font(bold=True, color=forest, size=10)
        for j in range(1, len(spans) + 2):
            ws.cell(row=r, column=j).fill = block_fill
        r += 1
        cells_by_line = []
        for ln in lns:
            vals = {(s, e): v for s, e, v in ln.entries}
            row_vals = [vals.get(sp) for sp in spans]
            cells_by_line.append(row_vals)
            ws.cell(row=r, column=1, value=f"  {ln.member}")
            ws.cell(row=r, column=1).font = Font(color=ink, size=10)
            for j, (sp, v) in enumerate(zip(spans, row_vals), start=2):
                cell = ws.cell(row=r, column=j)
                synth = sp in ln.synth
                if v is not None:
                    cell.value = v / _MM
                    cell.number_format = _FMT_MONEY
                cell.font = Font(color=ink, size=10, italic=synth)
                cell.alignment = Alignment(horizontal="right")
            r += 1
        if group != "Revenue":
            continue
        sums: List[Optional[float]] = []
        for i in range(len(spans)):
            vals = [rv[i] for rv in cells_by_line if rv[i] is not None]
            sums.append(sum(vals) if vals else None)
        cons_vals = [_find_span(cons, s, e) for s, e in spans]
        # FIX-14d: gate on the latest span where members have values
        # (annual preferred) — same partial-disclosure rule as the Model
        # sheet and the Phase-2 gate
        with_data = [i for i, sv in enumerate(sums) if sv is not None]
        annual = [i for i in with_data
                  if 330 <= (spans[i][1] - spans[i][0]).days <= 400]
        gate_i = max(annual) if annual else (
            max(with_data) if with_data else None)
        n_mem = 0 if gate_i is None else sum(
            1 for rv in cells_by_line if rv[gate_i] is not None)
        g_sigma = sums[gate_i] if gate_i is not None else None
        g_total = cons_vals[gate_i] if gate_i is not None else None
        if partial_axis_disclosure(n_mem, g_sigma, g_total):
            r = _write_note_row(ws, r, _tie_note(n_mem, g_sigma, g_total))
            continue
        # same check-row helper as the Model sheet (FIX-11b hoist) — one
        # styling for every Σ/gap tie row in the workbook
        r = _write_check_row(ws, r, "   Σ members", sums, ltm_col=0)
        r = _write_check_row(ws, r, "   vs consolidated (gap %)",
                             [_pct(sv, cv) for sv, cv in zip(sums, cons_vals)],
                             ltm_col=0, pct_fmt=True,
                             tol=config.SEGMENT_TIE_TOL)  # FIX-10d

    foot = [f"Source: {getattr(seg, 'source', '') or 'latest filings'} · "
            "columns are the fiscal spans as reported; italic cells were "
            "synthesized from the two-axis disaggregation table (not "
            "directly filed).",
            "Σ members / gap rows tie each Revenue axis to the consolidated "
            "statement — a positive gap beyond +2% signals hierarchical "
            "members double-counting; a negative gap means untagged members."]
    status = getattr(seg, "status", "")
    if status:
        foot.append(f"Status: {status}.")
    for attr, label in (("coverage", "Coverage"), ("recast_log", "Recasts"),
                        ("breaks", "Membership breaks")):
        extra = getattr(seg, attr, None)  # FIX-10 fields, when merged
        if extra:
            foot.append(f"{label}: {extra}")
    r += 1
    for note in foot:
        ws.cell(row=r, column=1, value=note).font = Font(color=muted, size=8)
        r += 1
    ws.column_dimensions["A"].width = 36
    for j in range(2, len(spans) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 13.5
