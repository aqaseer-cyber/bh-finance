"""Phase-5 verdict mechanics (mirroring the workbook) and the XLSX exporter."""
import datetime as dt

import pytest

from forensic_viz.dashboard import render_verdict
from forensic_viz.metrics import DashboardData
from forensic_viz.valuation import CaseInputs, ValuationInputs, build_valuation
from forensic_viz.verdict import build_verdict
from forensic_viz.workbook import fill_workbook


def _data(price=100.0, fcf=500e6, sbc=50e6, interest=0.0):
    d = DashboardData(ticker="T", company="T Inc", subtitle="",
                      generated=dt.date(2026, 7, 3))
    d.last_close = price
    d.price_dates = [dt.date(2026, 7, 1)]
    d.price_closes = [price]
    d.diluted_shares = [100e6]
    d.fcf = [fcf]
    d.fcf_ex_sbc = [fcf - sbc]
    d.effective_tax_rate = 0.21
    d.interest_expense = [interest if interest else None]
    d.fcff = [fcf + interest * (1 - 0.21)]
    d.sbc = [sbc]
    d.total_debt = [1e9]
    d.cash = [4e8]
    d.book_equity = [2e9]
    d.net_income = [3e8]
    d.sic_code = "3571"
    return d


def _dcf_res(d, rating_inputs=None):
    inputs = ValuationInputs(
        method="dcf", discount_rate=0.09,
        cases={"Bear": CaseInputs(g0=0.02, g_term=0.02),
               "Base": CaseInputs(g0=0.05, g_term=0.025),
               "Bull": CaseInputs(g0=0.09, g_term=0.03)})
    return inputs, build_valuation(d, inputs)


def test_verdict_dual_track_and_average():
    d = _data()
    inputs, res = _dcf_res(d)
    v = build_verdict(d, inputs, res)
    # Track A: Bear growths on as-reported FCFF; matches the Bear case FV
    assert v.fv_a == pytest.approx(res.cases[0].fv_ps)
    # Track B: Base growths on the ex-SBC base -> below the Base case FV
    assert v.fv_b < res.cases[1].fv_ps
    assert v.fv_avg == pytest.approx((v.fv_a + v.fv_b) / 2)
    assert v.mos == pytest.approx((v.fv_avg - 100.0) / 100.0)


def test_verdict_stress_reduces_fv():
    d = _data()
    inputs, res = _dcf_res(d)
    v = build_verdict(d, inputs, res)
    assert v.stressed_fv_avg < v.fv_avg  # -5% FCFF shock bites
    assert v.stressed_mos < v.mos
    assert "FCFF shock" in v.shock_label


def test_coherence_gate_mirrors_workbook():
    # Price far above FV -> deeply negative MoS; Hold must trip the gate
    d = _data(price=400.0)
    inputs, res = _dcf_res(d)
    v = build_verdict(d, inputs, res, rating="Hold")
    assert v.mos < -0.15
    assert v.coherence.startswith("CHECK")
    # Sell is coherent with a deeply negative MoS
    v2 = build_verdict(d, inputs, res, rating="Sell")
    assert v2.coherence == "ok"
    # No rating -> gate reports, never guesses (judgment stays with the user)
    v3 = build_verdict(d, inputs, res)
    assert v3.coherence == "no rating"


def test_coherence_gate_optionality_exception():
    d = _data(price=400.0)
    inputs, res = _dcf_res(d)
    assert res.implied_g is not None and res.implied_g > 0.035  # optionality priced
    v = build_verdict(d, inputs, res, rating="Hold",
                      optionality="pipeline drug X, priced by the market")
    assert v.coherence.startswith("ok")


def test_render_verdict_page(tmp_path):
    d = _data()
    inputs, res = _dcf_res(d)
    v = build_verdict(d, inputs, res, rating="Buy")
    out = tmp_path / "verdict.png"
    fig = render_verdict(d, res, v, str(out))
    assert out.exists() and out.stat().st_size > 25_000
    assert len(fig.axes) >= 3


def test_shell_has_exactly_133_blue_input_cells():
    """FIX-8: the shell's 133-blue-cell contract is load-bearing (workbook.py
    docstring, preflight lint). Guard it so a future shell swap fails loudly."""
    import openpyxl

    from forensic_viz.workbook import TEMPLATE
    wb = openpyxl.load_workbook(str(TEMPLATE))
    blue = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                col = c.font.color
                if col is not None and getattr(col, "rgb", None) \
                        and str(col.rgb).endswith("0000FF"):
                    blue += 1
    assert blue == 133


def test_workbook_fill_roundtrip(tmp_path, testco_facts):
    import openpyxl

    from forensic_viz.edgar import parse_companyfacts
    from forensic_viz.metrics import apply_track, build_fundamental_metrics
    d = DashboardData(ticker="TESTCO", company="TESTCO INC", subtitle="",
                      generated=dt.date(2026, 7, 3))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(parse_companyfacts(testco_facts, "TESTCO"), d)
    d.last_close = 100.0
    d.price_closes = [80.0, 100.0]
    d.price_dates = [dt.date(2026, 6, 30), dt.date(2026, 7, 1)]
    d.thesis = "A thesis."
    d.terminal_risk = "A terminal risk."

    inputs, res = _dcf_res(d)
    v = build_verdict(d, inputs, res, rating="Buy")
    out = tmp_path / "filled.xlsx"
    report = fill_workbook(d, str(out), res=res, verdict=v)
    assert report.filled >= 30
    assert report.analyst_cells  # the judgment to-do list is returned

    wb = openpyxl.load_workbook(str(out))
    assert wb["Control"]["B7"].value == "TESTCO"
    assert wb["Control"]["B9"].value == "Standard"
    assert wb["Control"]["B20"].value == "Buy"
    assert wb["Phase1_Anchor"]["B5"].value == 100.0
    # $mm scaling: latest revenue 1.9e9 -> COGS 60% = 1140 ($mm)
    assert wb["Phase2_UnitEcon"]["B20"].value == pytest.approx(1140.0)
    assert wb["Phase1_Anchor"]["A27"].value == "A thesis."
    assert wb["Phase2_UnitEcon"]["A42"].value == "A terminal risk."
    assert wb["FCFF_DCF"]["B9"].value == pytest.approx(0.02)   # Track A g0 = Bear
    assert wb["FCFF_DCF"]["C9"].value == pytest.approx(0.05)   # Track B g0 = Base
    assert wb["Phase5_Verdict"]["B33"].value == "Buy"
    # formulas untouched (the 360-formula contract stays intact)
    assert str(wb["Control"]["B19"].value).startswith("=")