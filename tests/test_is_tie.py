"""FIX-11b: the Rev − COGS vs GP tie row on the model export's face."""
import datetime as dt

import pytest
from openpyxl import load_workbook

from conftest import FY_YEARS, REVENUE, build_testco_companyfacts, _annual, _usd
from forensic_viz import config
from forensic_viz.edgar import parse_companyfacts
from forensic_viz.metrics import DashboardData, apply_track, build_fundamental_metrics
from forensic_viz.model_export import export_financial_model

TIE_LABEL = "   Rev − COGS vs GP (gap)"


def _facts_with_gp(offset: float) -> dict:
    facts = build_testco_companyfacts()
    facts["facts"]["us-gaap"]["GrossProfit"] = _usd(
        [_annual(y, REVENUE[y] * 0.4 + offset) for y in FY_YEARS])
    return facts


def _export(facts, tmp_path):
    d = DashboardData(ticker="T", company="T Inc", subtitle="",
                      generated=dt.date(2026, 5, 1))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(parse_companyfacts(facts, "T"), d)
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    ws = load_workbook(str(out))["Financial Model"]
    header = [c.value for c in ws[1]]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    return ws, header, labels


def test_coherent_statement_ties_at_zero_without_footnote(tmp_path):
    ws, header, labels = _export(_facts_with_gp(0.0), tmp_path)
    assert TIE_LABEL in labels
    row = labels.index(TIE_LABEL) + 1
    fy25 = header.index("FY2025") + 1
    assert ws.cell(row=row, column=fy25).value == pytest.approx(0.0)
    joined = " ".join(str(v) for v in labels if v)
    assert "Income-statement basis check" not in joined


def test_basis_break_flags_cell_and_footnote(tmp_path):
    # +200e6 on GP: no revenue candidate is coherent (11a leaves an
    # UNRESOLVED note) and the sheet's tie row must expose the breach
    ws, header, labels = _export(_facts_with_gp(200e6), tmp_path)
    row = labels.index(TIE_LABEL) + 1
    fy25 = header.index("FY2025") + 1
    cell = ws.cell(row=row, column=fy25)
    expected = -200e6 / REVENUE[2025]
    assert cell.value == pytest.approx(expected)
    assert abs(cell.value) > config.IS_TIE_TOL
    bad = "B3402A"
    assert str(cell.font.color.rgb).endswith(bad)  # flag-red beyond tol
    joined = " ".join(str(v) for v in labels if v)
    assert "Income-statement basis check" in joined
    assert "not on one accounting basis" in joined
