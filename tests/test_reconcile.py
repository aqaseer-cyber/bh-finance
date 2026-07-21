"""FIX-17c: reconciliation + gap-rescue audit — hand-computed outcomes,
sign conventions, the zero-is-absent rule, Finnhub tag-priority lookup,
graceful degradation, and the export block. All offline."""
import datetime as dt
from types import SimpleNamespace

import pytest

from forensic_viz import config
from forensic_viz.metrics import DashboardData
from forensic_viz.reconcile import (
    AuditEntry, AuditReport, fmt_val, reconcile_finnhub, reconcile_fmp,
    run_reconciliation,
)

FY23, FY24 = dt.date(2023, 12, 31), dt.date(2024, 12, 31)


def _d(series):
    d = DashboardData(ticker="T", company="T Inc", subtitle="",
                      generated=dt.date(2026, 7, 18))
    d.fundamentals = SimpleNamespace(fy_ends=[FY23, FY24], series=series)
    return d


def test_fmp_reconcile_hand_computed():
    d = _d({
        "revenue": [90e6, 100e6],
        "capex": [None, 500e6],          # positive outflow, our basis
        "sbc": [None, None],             # EDGAR never tagged it
        "dividends_paid": [10e6, 12e6],
        "diluted_shares": [100e6, 100e6],
    })
    statements = {
        "income": [
            {"date": "2024-12-31", "revenue": 101e6,       # within 2%
             "weightedAverageShsOutDil": 103.1e6},          # > 2% -> flag
            {"date": "2023-12-31", "revenue": 96e6},        # > tolerance
        ],
        "cashflow": [
            {"date": "2024-12-31",
             "capitalExpenditure": -510e6,   # FMP sign; |diff| in tol
             "stockBasedCompensation": 50e6,  # ours None -> rescuable
             "dividendsPaid": 0.0},           # zero = absent, no entry
        ],
        "balance": [],
    }
    rep = AuditReport()
    reconcile_fmp(d, statements, rep)
    # checked pairs: revenue FY24, shares FY24, revenue FY23, capex FY24
    assert rep.checked == 4
    assert rep.matched == 2                      # revenue FY24 + capex
    kinds = {(e.item, e.fy): e.kind for e in rep.entries}
    assert kinds[("Diluted shares", "FY2024")] == "divergent"
    assert kinds[("Revenue", "FY2023")] == "divergent"
    assert kinds[("SBC", "FY2024")] == "rescuable"
    assert ("Dividends paid", "FY2024") not in kinds
    resc = next(e for e in rep.entries if e.kind == "rescuable")
    assert resc.theirs == pytest.approx(50e6)
    assert resc.ours is None and resc.source == "FMP"


def test_finnhub_lookup_respects_tag_priority_and_prefix():
    d = _d({"revenue": [None, 100e6], "operating_income": [None, None],
            "net_income": [None, 20e6], "cfo": [None, 30e6]})
    payload = {"data": [{
        "endDate": "2024-12-31 00:00:00",
        "report": {
            "ic": [
                # extension tag: must be ignored
                {"concept": "meli_TotalMadeUpRevenue", "value": 999e6},
                # lower-priority us-gaap tag present…
                {"concept": "us-gaap_Revenues", "value": 100.5e6},
                # …and the higher-priority one wins
                {"concept":
                 "us-gaap_RevenueFromContractWithCustomer"
                 "ExcludingAssessedTax", "value": 100.4e6},
                {"concept": "us-gaap:NetIncomeLoss", "value": 27e6},
            ],
            "cf": [{"concept": "us-gaap_NetCashProvidedByUsedIn"
                               "OperatingActivities", "value": 30.2e6}],
        },
    }]}
    rep = AuditReport()
    reconcile_finnhub(d, payload, rep)
    by_item = {e.item: e for e in rep.entries}
    # revenue matched (100.4 vs 100 within 2%), cfo matched,
    # net income divergent (27 vs 20; as-first-filed comparison)
    assert rep.matched == 2
    assert by_item["Net income"].kind == "divergent"
    assert by_item["Net income"].theirs == pytest.approx(27e6)
    assert by_item["Net income"].source == "FNH"
    # operating income absent on both sides -> no entry, not checked
    assert "Operating income" not in by_item


def test_run_reconciliation_degrades_never_raises(monkeypatch):
    import forensic_viz.providers.finnhub as fnh
    import forensic_viz.providers.fmp as fmp
    monkeypatch.setattr(config, "FMP_API_KEY", "k")
    monkeypatch.setattr(config, "FINNHUB_API_KEY", "k")

    class Boom:
        def __init__(self, *a, **kw):
            pass

        def __getattr__(self, name):
            def die(*a, **kw):
                raise RuntimeError("proxy down")
            return die

    monkeypatch.setattr(fmp, "FMPClient", Boom)
    monkeypatch.setattr(fnh, "FinnhubClient", Boom)
    rep = run_reconciliation(_d({"revenue": [1e6, 2e6]}), cache=None)
    assert rep.checked == 0 and rep.entries == []
    assert "FMP" in rep.error and "Finnhub" in rep.error
    assert "unavailable" in rep.summary()


def test_summary_and_fmt():
    rep = AuditReport(checked=10, matched=9, sources=["FMP"])
    rep.entries.append(AuditEntry("Revenue", "FY2024", 1e9, 2e9,
                                  "FMP", "divergent"))
    s = rep.summary()
    assert "10 item-years" in s and "9 match" in s and "1 divergent" in s
    assert fmt_val(None, "money") == "–"
    assert fmt_val(2.5e9, "money") == "$2,500M"
    assert fmt_val(120e6, "shares") == "120.0M sh"


def test_export_carries_the_data_audit_block(tmp_path):
    from openpyxl import load_workbook

    from forensic_viz.edgar import parse_companyfacts
    from forensic_viz.metrics import apply_track, build_fundamental_metrics
    from forensic_viz.model_export import export_financial_model
    from test_model_export import _facts_with_quarters

    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 8, 10))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(
        parse_companyfacts(_facts_with_quarters(), "TESTCO"), d)
    rep = AuditReport(checked=8, matched=7, sources=["FMP"])
    rep.entries.append(AuditEntry("Revenue", "FY2024", 90e6, 100e6,
                                  "FMP", "divergent"))
    rep.entries.append(AuditEntry("SBC", "FY2025", None, 50e6,
                                  "FMP", "rescuable"))
    d.audit_report = rep

    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    wb = load_workbook(str(out))
    # v3 R3c: the audit table is a real table on the Audit sheet; the
    # Model sheet keeps a one-line pointer
    fm_a = [str(wb["Financial Model"].cell(row=r, column=1).value or "")
            for r in range(1, wb["Financial Model"].max_row + 1)]
    assert not any(v.startswith("DATA AUDIT") for v in fm_a)
    assert any("Audit sheet carries" in v for v in fm_a)
    ws = wb["Audit"]
    col_a = [str(ws.cell(row=r, column=1).value or "")
             for r in range(1, ws.max_row + 1)]
    assert any("8 item-years" in v for v in col_a)
    rows = [[ws.cell(row=r, column=c).value for c in range(1, 7)]
            for r in range(1, ws.max_row + 1)]
    flat = ["|".join(str(x) for x in row) for row in rows]
    assert any("Revenue|FY2024|$90M|$100M|FMP (aggregator)|DIVERGENT"
               in f for f in flat)
    assert any("RESCUABLE (EDGAR empty)" in f for f in flat)
    assert any(v.startswith("Tolerance ±2%") for v in col_a)


def test_export_without_audit_has_no_block(tmp_path):
    from openpyxl import load_workbook

    from forensic_viz.edgar import parse_companyfacts
    from forensic_viz.metrics import apply_track, build_fundamental_metrics
    from forensic_viz.model_export import export_financial_model
    from test_model_export import _facts_with_quarters

    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 8, 10))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(
        parse_companyfacts(_facts_with_quarters(), "TESTCO"), d)
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    wb = load_workbook(str(out))
    for sheet in ("Financial Model", "Audit"):
        col_a = [str(wb[sheet].cell(row=r, column=1).value or "")
                 for r in range(1, wb[sheet].max_row + 1)]
        if sheet == "Financial Model":
            assert not any(v.startswith("DATA AUDIT") for v in col_a)
        else:  # the absence is DECLARED on the Audit sheet
            assert any("provider recheck off" in v for v in col_a)
