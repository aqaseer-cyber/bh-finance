"""Segment extraction from XBRL instances + SEGMENTS section + Phase-2 fill.

The synthetic instance mirrors MELI's structure: revenue disaggregated on
the geography axis (Brazil / Mexico / Other Countries) and the
product/service axis (Commerce / Fintech Services), with operating income
on the product axis. Includes rejection cases: an instant context, a
two-axis context, a non-USD unit, and a neutral OperatingSegments marker
that must NOT disqualify a context.
"""
import datetime as dt

import pytest
from openpyxl import load_workbook

from conftest import REVENUE, build_testco_companyfacts
from forensic_viz.edgar import instance_url, parse_companyfacts
from forensic_viz.metrics import DashboardData, apply_track, build_fundamental_metrics
from forensic_viz.model_export import export_financial_model
from forensic_viz.segments import (
    SegmentData, build_segment_data, member_label, parse_instance,
)
from forensic_viz.workbook import fill_workbook

_NS = ('xmlns="http://www.xbrl.org/2003/instance" '
       'xmlns:xbrldi="http://xbrl.org/2006/xbrldi" '
       'xmlns:us-gaap="http://fasb.org/us-gaap/2025" '
       'xmlns:srt="http://fasb.org/srt/2025" '
       'xmlns:meli="http://mercadolibre.com/20251231" '
       'xmlns:iso4217="http://www.xbrl.org/2003/iso4217"')

REV = "us-gaap:RevenueFromContractWithCustomerExcludingAssessedTax"


def _ctx(cid, axis, member, start, end, extra=""):
    return f"""
 <context id="{cid}"><entity><identifier scheme="http://www.sec.gov/CIK">0001099590</identifier>
  <segment><xbrldi:explicitMember dimension="{axis}">{member}</xbrldi:explicitMember>{extra}</segment></entity>
  <period><startDate>{start}</startDate><endDate>{end}</endDate></period></context>"""


def _fact(concept, ctx, val, unit="usd"):
    return f'\n <{concept} contextRef="{ctx}" unitRef="{unit}" decimals="-5">{val}</{concept}>'


def _tenk_instance() -> str:
    """FY2024+FY2025 segment revenue (geo + product), op income (product)."""
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>',
             '<unit id="brl"><measure>iso4217:BRL</measure></unit>']
    geo = {"meli:BrazilMember": (900e6, 1000e6),
           "country:MX": (500e6, 560e6),
           "meli:OtherCountriesMember": (300e6, 340e6)}
    for i, (m, (v24, v25)) in enumerate(geo.items()):
        parts.append(_ctx(f"g24{i}", "srt:StatementGeographicalAxis", m,
                          "2024-01-01", "2024-12-31"))
        parts.append(_ctx(f"g25{i}", "srt:StatementGeographicalAxis", m,
                          "2025-01-01", "2025-12-31"))
        parts.append(_fact(REV, f"g24{i}", v24))
        parts.append(_fact(REV, f"g25{i}", v25))
    prod = {"meli:CommerceMember": (1000e6, 1080e6),
            "meli:FintechServicesMember": (700e6, 820e6)}
    for i, (m, (v24, v25)) in enumerate(prod.items()):
        # neutral OperatingSegments marker must not disqualify the context
        extra = ('<xbrldi:explicitMember dimension="us-gaap:ConsolidationItemsAxis">'
                 'us-gaap:OperatingSegmentsMember</xbrldi:explicitMember>')
        parts.append(_ctx(f"p24{i}", "srt:ProductOrServiceAxis", m,
                          "2024-01-01", "2024-12-31", extra))
        parts.append(_ctx(f"p25{i}", "srt:ProductOrServiceAxis", m,
                          "2025-01-01", "2025-12-31", extra))
        parts.append(_fact(REV, f"p24{i}", v24))
        parts.append(_fact(REV, f"p25{i}", v25))
        parts.append(_fact("us-gaap:OperatingIncomeLoss", f"p25{i}", v25 * 0.2))
    # rejects: two-axis context, non-USD fact, sub-20-day span
    parts.append(_ctx("bad2ax", "srt:StatementGeographicalAxis",
                      "meli:BrazilMember", "2025-01-01", "2025-12-31",
                      '<xbrldi:explicitMember dimension="srt:ProductOrServiceAxis">'
                      'meli:CommerceMember</xbrldi:explicitMember>'))
    parts.append(_fact(REV, "bad2ax", 999e9))
    parts.append(_ctx("badbrl", "srt:StatementGeographicalAxis",
                      "meli:BrazilMember", "2025-01-01", "2025-12-31"))
    parts.append(_fact(REV, "badbrl", 888e9, unit="brl"))
    parts.append("</xbrl>")
    return "".join(parts)


def _tenq_instance() -> str:
    """Q1'26 discrete + prior-year Q1'25 comparative for the geo axis."""
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>']
    q = {"meli:BrazilMember": (230e6, 270e6),
         "country:MX": (130e6, 150e6),
         "meli:OtherCountriesMember": (80e6, 90e6)}
    for i, (m, (v25, v26)) in enumerate(q.items()):
        parts.append(_ctx(f"q25{i}", "srt:StatementGeographicalAxis", m,
                          "2025-01-01", "2025-03-31"))
        parts.append(_ctx(f"q26{i}", "srt:StatementGeographicalAxis", m,
                          "2026-01-01", "2026-03-31"))
        parts.append(_fact(REV, f"q25{i}", v25))
        parts.append(_fact(REV, f"q26{i}", v26))
    parts.append("</xbrl>")
    return "".join(parts)


def test_member_label_and_instance_url():
    assert member_label("meli:FintechServicesMember") == "Fintech Services"
    assert member_label("country:MX") == "Mexico"  # ISO code -> country name
    assert member_label("country:BR") == "Brazil"
    # FIX-13c: trailing 'Segment' strips too, so X and XSegment merge
    assert member_label("meli:BrazilSegmentMember") == "Brazil"
    assert member_label("meli:OtherCountriesSegmentMember") == \
        member_label("meli:OtherCountriesMember") == "Other Countries"
    assert instance_url(1099590, "0001099590-26-000012", "meli-20251231.htm") \
        == ("https://www.sec.gov/Archives/edgar/data/1099590/"
            "000109959026000012/meli-20251231_htm.xml")


def test_parse_instance_accepts_segment_axes_and_rejects_noise():
    parsed = parse_instance(_tenk_instance())
    rev_local = REV.split(":")[1]
    br = parsed.singles[("geography", "Brazil", rev_local)]
    assert br[(dt.date(2025, 1, 1), dt.date(2025, 12, 31))] == 1000e6
    assert all(v < 1e11 for v in br.values())  # BRL-unit fact rejected
    # the two-axis (geo × product) fact lands in crosses, not in singles
    assert any(k[-1] == rev_local and v.get(
        (dt.date(2025, 1, 1), dt.date(2025, 12, 31))) == 999e9
        for k, v in parsed.crosses.items())
    # neutral OperatingSegments marker kept the product contexts alive
    assert ("product / service", "Fintech Services", rev_local) \
        in parsed.singles
    assert ("product / service", "Commerce", "OperatingIncomeLoss") \
        in parsed.singles


def test_build_orders_members_by_revenue_and_merges_instances():
    seg = build_segment_data([("10-K a25 (FY2025)", _tenk_instance()),
                              ("10-Q q26", _tenq_instance())], "test")
    assert seg.axes() == ["product / service", "geography"]
    assert seg.members("geography") == ["Brazil", "Mexico", "Other Countries"]
    assert seg.n_segments == 2  # primary axis = product / service
    rev_br = next(ln for ln in seg.lines
                  if ln.member == "Brazil" and ln.group == "Revenue")
    spans = {(s, e): v for s, e, v in rev_br.entries}
    assert spans[(dt.date(2026, 1, 1), dt.date(2026, 3, 31))] == 270e6
    # the noisy two-axis 999e9 fact never overrode the filed Brazil total
    assert spans[(dt.date(2025, 1, 1), dt.date(2025, 12, 31))] == 1000e6


def test_two_axis_only_disaggregation_is_aggregated():
    """MELI-style: the disaggregation table is tagged geography × business
    with NO single-axis totals — totals must be synthesized per axis."""
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>']
    cells = {("country:BR", "meli:CommerceMember"): 600e6,
             ("country:BR", "meli:FintechMember"): 400e6,
             ("country:MX", "meli:CommerceMember"): 250e6,
             ("country:MX", "meli:FintechMember"): 150e6}
    for i, ((geo, biz), v) in enumerate(cells.items()):
        extra = ('<xbrldi:explicitMember dimension="srt:ProductOrServiceAxis">'
                 f"{biz}</xbrldi:explicitMember>")
        parts.append(_ctx(f"x{i}", "srt:StatementGeographicalAxis", geo,
                          "2025-01-01", "2025-12-31", extra))
        parts.append(_fact(REV, f"x{i}", v))
    parts.append("</xbrl>")
    seg = build_segment_data([("10-K a25 (FY2025)", "".join(parts))], "x")
    by = {(ln.axis, ln.member): ln for ln in seg.lines
          if ln.group == "Revenue"}
    span = (dt.date(2025, 1, 1), dt.date(2025, 12, 31))
    assert {(s, e): v for s, e, v in
            by[("geography", "Brazil")].entries}[span] == 1000e6
    assert {(s, e): v for s, e, v in
            by[("product / service", "Fintech")].entries}[span] == 550e6
    assert "aggregated" in seg.status


def test_subsegments_axis_accepted_as_revenue_stream():
    """FIX-13b, verified on MELI's FY2024 instance: the Commerce/Fintech
    split rides srt:SubsegmentsAxis (Commerce 12,159 / Fintech 8,618 $mm,
    Σ = 20,777), with Country×Stream crosses that tie per country
    (Brazil 7,038 + 4,368 = 11,406)."""
    span = ("2024-01-01", "2024-12-31")
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>']
    # stream singles tagged directly, as MELI files them
    streams = {"meli:CommerceMember": 12_159e6,
               "meli:FintechMember": 8_618e6}
    for i, (m, v) in enumerate(streams.items()):
        parts.append(_ctx(f"ss{i}", "srt:SubsegmentsAxis", m, *span))
        parts.append(_fact(REV, f"ss{i}", v))
    # Country×Stream crosses (ordinary 2-dim facts)
    crosses = {("country:BR", "meli:CommerceMember"): 7_038e6,
               ("country:BR", "meli:FintechMember"): 4_368e6,
               ("country:MX", "meli:CommerceMember"): 2_800e6,
               ("country:MX", "meli:FintechMember"): 1_864e6,
               ("meli:OtherCountriesMember", "meli:CommerceMember"): 2_321e6,
               ("meli:OtherCountriesMember", "meli:FintechMember"): 2_386e6}
    for i, ((geo, st), v) in enumerate(crosses.items()):
        extra = ('<xbrldi:explicitMember dimension="srt:SubsegmentsAxis">'
                 f"{st}</xbrldi:explicitMember>")
        parts.append(_ctx(f"cx{i}", "srt:StatementGeographicalAxis", geo,
                          *span, extra))
        parts.append(_fact(REV, f"cx{i}", v))
    parts.append("</xbrl>")
    seg = build_segment_data([("10-K m24 (FY2024)", "".join(parts))],
                             "meli-style")

    assert "revenue stream" in seg.axes()
    # axis ordering follows _AXIS_ORDER: stream before geography
    assert seg.axes().index("revenue stream") < seg.axes().index("geography")
    by = {(ln.axis, ln.member): ln for ln in seg.lines
          if ln.group == "Revenue"}
    dspan = (dt.date(2024, 1, 1), dt.date(2024, 12, 31))

    def val(axis, member):
        return {(s, e): v for s, e, v in by[(axis, member)].entries}[dspan]

    assert val("revenue stream", "Commerce") == 12_159e6   # filed single wins
    assert val("revenue stream", "Fintech") == 8_618e6
    total = 12_159e6 + 8_618e6                              # 20,777
    # per-span ties: Σ(stream) == Σ(country) == consolidated
    stream_sum = sum(val("revenue stream", m)
                     for m in seg.members("revenue stream"))
    geo_sum = sum(val("geography", m) for m in seg.members("geography"))
    assert stream_sum == pytest.approx(total)
    assert geo_sum == pytest.approx(total)
    assert val("geography", "Brazil") == pytest.approx(7_038e6 + 4_368e6)


def test_duplicate_member_qnames_merge_to_one_line():
    """FIX-13c, verified on MELI FY2024: the business-segments axis files
    OtherCountriesMember AND OtherCountriesSegmentMember (889 each) —
    pre-fix Σ(countries) inflated to 21,666 (+4.3%) and falsely tripped
    the tie gate."""
    span = ("2024-01-01", "2024-12-31")
    vals = {"meli:BrazilSegmentMember": 11_406e6,
            "meli:MexicoSegmentMember": 4_664e6,
            "meli:ArgentinaSegmentMember": 3_818e6}
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>']
    for i, (m, v) in enumerate(vals.items()):
        parts.append(_ctx(f"b{i}", "us-gaap:StatementBusinessSegmentsAxis",
                          m, *span))
        parts.append(_fact(REV, f"b{i}", v))
    for i, m in enumerate(("meli:OtherCountriesMember",
                           "meli:OtherCountriesSegmentMember")):
        parts.append(_ctx(f"o{i}", "us-gaap:StatementBusinessSegmentsAxis",
                          m, *span))
        parts.append(_fact(REV, f"o{i}", 889e6))
    parts.append("</xbrl>")
    seg = build_segment_data([("10-K m24 (FY2024)", "".join(parts))],
                             "meli-style")

    members = seg.members("business segments")
    assert members.count("Other Countries") == 1      # one line, not two
    assert "Brazil" in members and "Brazil Segment" not in members
    dspan = (dt.date(2024, 1, 1), dt.date(2024, 12, 31))
    by = {ln.member: {(s, e): v for s, e, v in ln.entries}
          for ln in seg.lines if ln.group == "Revenue"}
    assert by["Other Countries"][dspan] == 889e6      # kept once, not summed
    # Σ(countries) ties back to the consolidated 20,777 (gap 0.0%)
    assert sum(v[dspan] for v in by.values()) == pytest.approx(20_777e6)
    assert "member aliases merged: Other Countries (2 qnames)" in seg.status


def test_disagreeing_qname_aliases_keep_first_and_warn():
    """FIX-13c conflict guard: > 1% disagreement between qname aliases on
    one span keeps the first value and warns — never averages."""
    span = ("2024-01-01", "2024-12-31")
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>']
    for i, (m, v) in enumerate((("meli:OtherCountriesMember", 889e6),
                                ("meli:OtherCountriesSegmentMember", 950e6))):
        parts.append(_ctx(f"o{i}", "us-gaap:StatementBusinessSegmentsAxis",
                          m, *span))
        parts.append(_fact(REV, f"o{i}", v))
    parts.append("</xbrl>")
    seg = build_segment_data([("10-K x24 (FY2024)", "".join(parts))], "x")
    dspan = (dt.date(2024, 1, 1), dt.date(2024, 12, 31))
    ln = next(x for x in seg.lines if x.member == "Other Countries")
    assert {(s, e): v for s, e, v in ln.entries}[dspan] == 889e6  # first kept
    assert "kept the first" in seg.status
    assert "member aliases merged" in seg.status  # both notes coexist


def test_three_axis_facts_counted_not_modeled():
    """FIX-13b: a Country×Stream×Product fact (all accepted axes) is not a
    line — it is counted and declared in the status."""
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>']
    extra = ('<xbrldi:explicitMember dimension="srt:SubsegmentsAxis">'
             'meli:CommerceMember</xbrldi:explicitMember>'
             '<xbrldi:explicitMember dimension="srt:ProductOrServiceAxis">'
             'meli:MarketplaceMember</xbrldi:explicitMember>')
    parts.append(_ctx("m3", "srt:StatementGeographicalAxis", "country:BR",
                      "2024-01-01", "2024-12-31", extra))
    parts.append(_fact(REV, "m3", 5_000e6))
    # plus one normal single so the parse has a line to keep
    parts.append(_ctx("s0", "srt:SubsegmentsAxis", "meli:CommerceMember",
                      "2024-01-01", "2024-12-31"))
    parts.append(_fact(REV, "s0", 12_159e6))
    parts.append("</xbrl>")
    parsed = parse_instance("".join(parts))
    assert parsed.n_multi == 1
    assert all(k[1] != "Brazil" for k in parsed.singles)  # not modeled

    seg = build_segment_data([("10-K x24 (FY2024)", "".join(parts))], "x")
    assert "1 facts at 3+ segment axes ignored" in seg.status
    assert seg.members("revenue stream") == ["Commerce"]


def _data_with_segments():
    facts = build_testco_companyfacts()
    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 5, 1))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(parse_companyfacts(facts, "TESTCO"), d)
    d.segments = build_segment_data([("10-K a25 (FY2025)", _tenk_instance()),
                                     ("10-Q q26", _tenq_instance())], "t")
    return d


def test_segments_section_in_model_export(tmp_path):
    d = _data_with_segments()
    out = tmp_path / "model.xlsx"
    export_financial_model(d, str(out))
    ws = load_workbook(str(out))["Financial Model"]
    header = [c.value for c in ws[1]]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "SEGMENTS (as filed)" in labels
    assert "Revenue by geography" in labels
    assert "Revenue by product / service" in labels
    assert "Operating income by product / service" in labels
    br_row = labels.index("  Brazil") + 1
    fy25 = header.index("FY2025") + 1
    assert ws.cell(row=br_row, column=fy25).value == pytest.approx(1000.0)
    # % change row under the Brazil segment line: FY2025 YoY
    assert labels[br_row] == "   % change"
    assert ws.cell(row=br_row + 1, column=fy25).value == pytest.approx(
        1000e6 / 900e6 - 1)


def test_workbook_phase2_segment_fill(tmp_path):
    d = _data_with_segments()
    out = tmp_path / "wb.xlsx"
    report = fill_workbook(d, str(out))
    wb = load_workbook(str(out))
    ws = wb["Phase2_UnitEcon"]
    # top-2 product/service segments in $mm; remainder = total − top-2
    assert ws["B5"].value == pytest.approx(1080.0)
    assert ws["B6"].value == pytest.approx(820.0)
    assert ws["B7"].value == pytest.approx(REVENUE[2025] / 1e6 - 1080 - 820)
    assert "Commerce" in ws["B5"].comment.text
    assert "Fintech Services" in ws["B6"].comment.text
    # the segment row left the analyst to-do list
    assert not any(r[0] == "Phase2_UnitEcon" and r[1] == "B5:B7"
                   for r in report.analyst_cells)


def test_no_segments_keeps_workbook_and_export_working(tmp_path):
    facts = build_testco_companyfacts()
    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 5, 1))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(parse_companyfacts(facts, "TESTCO"), d)
    assert d.segments is None
    report = fill_workbook(d, str(tmp_path / "wb.xlsx"))
    assert any(r[0] == "Phase2_UnitEcon" and r[1] == "B5:B7"
               for r in report.analyst_cells)
    export_financial_model(d, str(tmp_path / "m.xlsx"))
    ws = load_workbook(str(tmp_path / "m.xlsx"))["Financial Model"]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "SEGMENTS (as filed)" not in labels
    assert any("no dimensional segment data" in str(v) for v in labels)


def _single_axis_instance(members: dict, total_span=("2025-01-01", "2025-12-31")):
    """One-axis product/service instance with the given member -> value map."""
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>']
    for i, (m, v) in enumerate(members.items()):
        parts.append(_ctx(f"s{i}", "srt:ProductOrServiceAxis", m,
                          *total_span))
        parts.append(_fact(REV, f"s{i}", v))
    parts.append("</xbrl>")
    return "".join(parts)


def test_hierarchical_members_skip_phase2_fill(tmp_path):
    """Parent + child on one axis: Σ members > consolidated revenue. The
    Phase-2 fill must be SKIPPED with a note — top-2-by-size would put a
    child inside its parent, double-count, and still tie at B8."""
    total = REVENUE[2025]
    xml = _single_axis_instance({
        "meli:CommerceMember": total * 0.60,
        "meli:CommerceServicesMember": total * 0.35,  # child of Commerce
        "meli:AdvertisingMember": total * 0.10,
    })
    facts = build_testco_companyfacts()
    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 5, 1))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(parse_companyfacts(facts, "TESTCO"), d)
    d.segments = build_segment_data([("10-K a25 (FY2025)", xml)], "t")
    report = fill_workbook(d, str(tmp_path / "wb.xlsx"))
    assert any("SKIPPED" in n and "hierarchical" in n
               for n in (report.notes or []))
    wb = load_workbook(str(tmp_path / "wb.xlsx"))
    assert wb["Phase2_UnitEcon"]["B5"].value == 6000  # shell placeholder kept
    assert any(r[0] == "Phase2_UnitEcon" and r[1] == "B5:B7"
               for r in report.analyst_cells)  # still on the analyst to-do


def test_incomplete_cross_table_flagged_by_tie_rows(tmp_path):
    """Cross-only table covering ~80% of consolidated revenue: synthesized
    totals are italic and the gap row shows the shortfall in red-flag
    territory."""
    total = REVENUE[2025]
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>']
    cells = {("country:BR", "meli:CommerceMember"): total * 0.42,
             ("country:BR", "meli:FintechMember"): total * 0.21,
             ("country:MX", "meli:CommerceMember"): total * 0.17}
    for i, ((geo, biz), v) in enumerate(cells.items()):
        extra = ('<xbrldi:explicitMember dimension="srt:ProductOrServiceAxis">'
                 f"{biz}</xbrldi:explicitMember>")
        parts.append(_ctx(f"x{i}", "srt:StatementGeographicalAxis", geo,
                          "2025-01-01", "2025-12-31", extra))
        parts.append(_fact(REV, f"x{i}", v))
    parts.append("</xbrl>")

    d = _data_with_segments()
    d.segments = build_segment_data([("10-K a25 (FY2025)", "".join(parts))], "t")
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    ws = load_workbook(str(out))["Financial Model"]
    header = [c.value for c in ws[1]]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    fy25 = header.index("FY2025") + 1
    # synthesized Brazil total renders italic
    br_row = labels.index("  Brazil") + 1
    assert ws.cell(row=br_row, column=fy25).font.italic
    # tie rows present; the FY2025 gap reads Σ/consolidated − 1 = −20%
    gap_row = labels.index("   vs consolidated (gap %)") + 1
    assert ws.cell(row=gap_row, column=fy25).value == pytest.approx(-0.20)
    sig_row = labels.index("   Σ members") + 1
    assert ws.cell(row=sig_row, column=fy25).value == pytest.approx(
        total * 0.80 / 1e6)


def test_tie_rows_show_zero_gap_when_members_are_complete(tmp_path):
    d = _data_with_segments()  # product members sum to 1900 = consolidated
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    ws = load_workbook(str(out))["Financial Model"]
    header = [c.value for c in ws[1]]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    fy25 = header.index("FY2025") + 1
    gap_row = labels.index("   vs consolidated (gap %)") + 1
    assert ws.cell(row=gap_row, column=fy25).value == pytest.approx(0.0)


# ---------------------------------------------------------------- FIX-14d

def _us_only_instance():
    """Deliberately partial geography note: US-only revenue, a sliver of
    consolidated (the MELI wolf-cry case)."""
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>',
             _ctx("us25", "srt:StatementGeographicalAxis", "country:US",
                  "2025-01-01", "2025-12-31"),
             _fact(REV, "us25", REVENUE[2025] * 0.002),
             "</xbrl>"]
    return "".join(parts)


def test_partial_axis_suppresses_tie_rows_on_both_sheets(tmp_path):
    d = _data_with_segments()
    d.segments = build_segment_data(
        [("10-K a25 (FY2025)", _us_only_instance())], "t")
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    wb = load_workbook(str(out))
    for sheet in ("Financial Model", "Segments"):
        labels = [wb[sheet].cell(row=r, column=1).value
                  for r in range(1, wb[sheet].max_row + 1)]
        assert "   Σ members" not in labels, sheet          # no red gap row
        assert "   vs consolidated (gap %)" not in labels, sheet
        note = next((l for l in labels
                     if l and "tie suppressed" in l), None)
        assert note is not None, sheet
        assert "partial disclosure axis" in note
        assert "1 member(s)" in note and "0% of consolidated" in note


def test_two_member_axis_keeps_tie_rows_even_at_low_coverage(tmp_path):
    # two members at 40% of consolidated: ≥2 members -> the tie renders
    # (and its −60% gap is a REAL finding, not a wolf-cry)
    total = REVENUE[2025]
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>',
             _ctx("us25", "srt:StatementGeographicalAxis", "country:US",
                  "2025-01-01", "2025-12-31"),
             _fact(REV, "us25", total * 0.30),
             _ctx("ca25", "srt:StatementGeographicalAxis", "country:CA",
                  "2025-01-01", "2025-12-31"),
             _fact(REV, "ca25", total * 0.10),
             "</xbrl>"]
    d = _data_with_segments()
    d.segments = build_segment_data([("10-K a25 (FY2025)", "".join(parts))], "t")
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    ws = load_workbook(str(out))["Financial Model"]
    header = [c.value for c in ws[1]]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    fy25 = header.index("FY2025") + 1
    gap_row = labels.index("   vs consolidated (gap %)") + 1
    assert ws.cell(row=gap_row, column=fy25).value == pytest.approx(-0.60)
    assert not any(l and "tie suppressed" in l for l in labels)


def test_partial_axis_ineligible_for_phase2_fill(tmp_path):
    """Members never share a FY span and the latest year is a US-only
    sliver: the Phase-2 gate reports ineligible-partial (quiet), not the
    fill-manually alarm."""
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>',
             _ctx("br24", "srt:StatementGeographicalAxis",
                  "meli:BrazilMember", "2024-01-01", "2024-12-31"),
             _fact(REV, "br24", 500e6),
             _ctx("us25", "srt:StatementGeographicalAxis", "country:US",
                  "2025-01-01", "2025-12-31"),
             _fact(REV, "us25", REVENUE[2025] * 0.002),
             "</xbrl>"]
    d = _data_with_segments()
    d.segments = build_segment_data([("10-K a25 (FY2025)", "".join(parts))], "t")
    report = fill_workbook(d, str(tmp_path / "wb.xlsx"))
    notes = report.notes or []
    assert any("ineligible for auto-fill" in n and "partial disclosure" in n
               for n in notes)
    assert not any("no fiscal-year span shared" in n for n in notes)
    wb = load_workbook(str(tmp_path / "wb.xlsx"))
    assert wb["Phase2_UnitEcon"]["B5"].value == 6000  # shell placeholder kept


# ---------------------------------------------------------------- FIX-10c

def _annual_instance(member_years, axis="srt:ProductOrServiceAxis", tag=REV):
    """{member qname: {year: value}} -> one-axis annual instance XML."""
    parts = [f"<xbrl {_NS}>",
             '<unit id="usd"><measure>iso4217:USD</measure></unit>']
    i = 0
    for m, years in member_years.items():
        for y, v in years.items():
            parts.append(_ctx(f"c{i}", axis, m, f"{y}-01-01", f"{y}-12-31"))
            parts.append(_fact(tag, f"c{i}", v))
            i += 1
    parts.append("</xbrl>")
    return "".join(parts)


def test_stitch_across_instances_latest_restated_wins_with_recast_log():
    i1 = _annual_instance({"meli:CommerceMember": {2021: 100e6, 2022: 110e6}})
    i2 = _annual_instance({"meli:CommerceMember": {2023: 120e6, 2024: 130e6}})
    i3 = _annual_instance({"meli:CommerceMember": {2024: 135e6,   # restated
                                                   2025: 150e6}})
    seg = build_segment_data([("10-K k21 (FY2022)", i1),
                              ("10-K k23 (FY2024)", i2),
                              ("10-K k25 (FY2025)", i3)])
    ln = next(l for l in seg.lines
              if l.member == "Commerce" and l.group == "Revenue")
    got = {e.year: v for s, e, v in ln.entries}
    assert got == {2021: 100e6, 2022: 110e6, 2023: 120e6,
                   2024: 135e6, 2025: 150e6}   # five years, newest FY24 wins
    assert len(seg.recast_log) == 1
    assert "10-K k23 (FY2024)" in seg.recast_log[0]
    assert "10-K k25 (FY2025)" in seg.recast_log[0]
    assert "135,000,000" in seg.recast_log[0]
    assert seg.source == "10-K FY2022–FY2025"  # derived from the labels


def test_rename_at_boundary_never_splices_without_an_alias():
    old = _annual_instance({"meli:MarketplaceMember":
                            {2021: 100e6, 2022: 110e6}})
    new = _annual_instance({"meli:CommerceMember":
                            {2022: 110e6, 2023: 120e6}})
    inst = [("10-K k22 (FY2022)", old), ("10-K k23 (FY2023)", new)]

    seg = build_segment_data(inst)
    members = {ln.member for ln in seg.lines if ln.group == "Revenue"}
    assert members == {"Marketplace", "Commerce"}   # two lines, no merge
    assert len(seg.breaks) == 1
    assert "retired ['Marketplace']" in seg.breaks[0]
    assert "introduced ['Commerce']" in seg.breaks[0]

    aliased = build_segment_data(inst, aliases={"Marketplace": "Commerce"})
    lines = [ln for ln in aliased.lines if ln.group == "Revenue"]
    assert [ln.member for ln in lines] == ["Commerce"]  # one spliced line
    assert {e.year for s, e, v in lines[0].entries} == {2021, 2022, 2023}
    assert aliased.breaks == []                     # identity was declared


def test_alias_collapse_within_one_instance_keeps_first_and_warns():
    """Bundle hardening: an alias folding two members that BOTH appear in
    one instance on the same span must not silently overwrite — keep the
    first, warn beyond 1% (the qname guard's semantics), so the alias map
    can be corrected instead of quietly rewriting history."""
    xml = _annual_instance({"meli:CommerceMember": {2024: 100e6},
                            "meli:CommerceServicesMember": {2024: 35e6}})
    seg = build_segment_data([("10-K k24 (FY2024)", xml)],
                             aliases={"Commerce Services": "Commerce"})
    ln = next(l for l in seg.lines
              if l.member == "Commerce" and l.group == "Revenue")
    assert {e.year: v for s, e, v in ln.entries} == {2024: 100e6}  # first kept
    assert "alias collapse disagreement for Commerce" in seg.status
    assert "check the [segment_aliases] map" in seg.status
    # an agreeing collapse (≤1%) merges without the warning
    xml2 = _annual_instance({"meli:CommerceMember": {2024: 100e6},
                             "meli:MarketplaceMember": {2024: 100e6}})
    seg2 = build_segment_data([("10-K k24 (FY2024)", xml2)],
                              aliases={"Marketplace": "Commerce"})
    assert "alias collapse disagreement" not in seg2.status


def test_coverage_counts_and_malformed_instance_skipped():
    good1 = _annual_instance({"meli:CommerceMember": {2024: 100e6}})
    good2 = _annual_instance({"meli:CommerceMember": {2025: 120e6}})
    seg = build_segment_data([("10-K a (FY2024)", good1),
                              ("10-K broken (FY2025)", "not xml <<<"),
                              ("10-K b (FY2025)", good2)])
    assert [c for c in seg.coverage] == [("10-K a (FY2024)", 1),
                                         ("10-K b (FY2025)", 1)]
    assert any("10-K broken (FY2025): parse error" in s
               for s in seg.status.split("skipped: ")[-1:])
    assert len(seg.lines) == 1 and not seg.lines[0].discontinuous


def test_interior_gap_with_axis_peers_marks_discontinuous():
    xml = _annual_instance({
        "meli:AdsMember": {2021: 50e6, 2023: 60e6},           # hole at 2022
        "meli:CommerceMember": {2021: 100e6, 2022: 110e6, 2023: 120e6},
    })
    seg = build_segment_data([("10-K k (FY2023)", xml)])
    by = {ln.member: ln for ln in seg.lines if ln.group == "Revenue"}
    assert by["Ads"].discontinuous is True
    assert by["Commerce"].discontinuous is False


# ---------------------------------------------------------------- FIX-10d

def test_retired_member_does_not_poison_the_same_span_gate(tmp_path):
    """A member whose last data is FY-1 must not mix fiscal years into
    sigma: the gate works on the last COMMON span, the fill proceeds."""
    total = REVENUE[2025]
    xml = _annual_instance({
        "meli:CommerceMember": {2024: 1000e6, 2025: total - 800e6},
        "meli:FintechMember": {2024: 850e6, 2025: 800e6},
        "meli:LegacyMember": {2024: 50e6},              # retired after FY24
    })
    facts = build_testco_companyfacts()
    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 5, 1))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(parse_companyfacts(facts, "TESTCO"), d)
    d.segments = build_segment_data([("10-K k (FY2025)", xml)])
    report = fill_workbook(d, str(tmp_path / "wb.xlsx"))
    assert not any("SKIPPED" in n for n in (report.notes or []))
    wb = load_workbook(str(tmp_path / "wb.xlsx"))
    ws = wb["Phase2_UnitEcon"]
    assert ws["B5"].value == pytest.approx((total - 800e6) / 1e6)  # FY25 top
    assert ws["B6"].value == pytest.approx(800.0)
    assert "FY2025" in ws["B5"].comment.text
    # old latest()-based gate would have summed 1100+800+50 > total*1.02


def test_audit_footnotes_carry_coverage_breaks_and_recasts(tmp_path):
    old = _annual_instance({"meli:MarketplaceMember":
                            {2023: 900e6, 2024: 950e6}})
    new = _annual_instance({"meli:CommerceMember":
                            {2024: 990e6, 2025: 1100e6},
                            "meli:FintechMember":
                            {2024: 850e6, 2025: 800e6}})
    d = _data_with_segments()
    d.segments = build_segment_data([("10-K k24 (FY2024)", old),
                                     ("10-K k25 (FY2025)", new)])
    assert d.segments.breaks       # membership changed at the shared FY24
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    ws = load_workbook(str(out))["Financial Model"]
    labels = [str(ws.cell(row=r, column=1).value or "")
              for r in range(1, ws.max_row + 1)]
    joined = " ".join(labels)
    assert "Segment coverage: dimensional facts found in 2/2 instances" \
        in joined
    assert "Segment recast — series are not comparable across this " \
           "boundary" in joined
    assert "retired ['Marketplace']" in joined
