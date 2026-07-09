"""FIX-13d: as-filed statement sheets — FilingSummary/presentation/label
linkbase parsing (pure, offline) and the Income/Balance/Cash-Flow/Segments
sheets in the financial-model export."""
import datetime as dt

import pytest
from openpyxl import load_workbook

from conftest import REVENUE, SHARES, build_testco_companyfacts
from forensic_viz.edgar import (
    PresRow, annual_values_for_concept, build_statement_rows,
    parse_companyfacts, parse_filing_summary,
)
from forensic_viz.metrics import DashboardData, apply_track, build_fundamental_metrics
from forensic_viz.model_export import (
    _FMT_MONEY, _FMT_SHARES, export_financial_model,
)

_XL = 'xmlns:link="http://www.xbrl.org/2003/linkbase" ' \
      'xmlns:xlink="http://www.w3.org/1999/xlink"'
_ROLE_IS = "http://testco.example/role/IncomeStatement"
_ROLE_BS = "http://testco.example/role/BalanceSheet"
_TOTAL = "http://www.xbrl.org/2003/role/totalLabel"
_NEGATED = "http://www.xbrl.org/2003/role/negatedLabel"
_STD = "http://www.xbrl.org/2003/role/label"

FILING_SUMMARY = f"""<FilingSummary><MyReports>
 <Report><ShortName>Cover Page</ShortName>
  <Role>http://testco.example/role/Cover</Role></Report>
 <Report><ShortName>CONSOLIDATED STATEMENTS OF INCOME</ShortName>
  <Role>{_ROLE_IS}</Role></Report>
 <Report><ShortName>CONSOLIDATED BALANCE SHEETS</ShortName>
  <Role>{_ROLE_BS}</Role></Report>
 <Report><ShortName>CONSOLIDATED BALANCE SHEETS (Parenthetical)</ShortName>
  <Role>http://testco.example/role/BSParen</Role></Report>
 <Report><ShortName>CONSOLIDATED STATEMENTS OF CASH FLOWS</ShortName>
  <Role>http://testco.example/role/CashFlow</Role></Report>
</MyReports></FilingSummary>"""


def _loc(label, concept):
    return (f'<link:loc xlink:label="{label}" '
            f'xlink:href="tc.xsd#us-gaap_{concept}"/>')


def _arc(frm, to, order, pref=""):
    p = f' preferredLabel="{pref}"' if pref else ""
    return (f'<link:presentationArc xlink:from="{frm}" xlink:to="{to}" '
            f'order="{order}"{p}/>')


PRE_XML = f"""<link:linkbase {_XL}>
 <link:presentationLink xlink:role="{_ROLE_IS}">
  {_loc("l_abs", "IncomeStatementAbstract")}
  {_loc("l_rev", "RevenueFromContractWithCustomerExcludingAssessedTax")}
  {_loc("l_cogs", "CostOfRevenue")}
  {_loc("l_gp", "GrossProfit")}
  {_loc("l_ni", "NetIncomeLoss")}
  {_arc("l_abs", "l_rev", 1)}
  {_arc("l_abs", "l_cogs", 2, _NEGATED)}
  {_arc("l_abs", "l_gp", 3, _TOTAL)}
  {_arc("l_abs", "l_ni", 4, _TOTAL)}
 </link:presentationLink>
 <link:presentationLink xlink:role="{_ROLE_BS}">
  {_loc("l_bs", "StatementOfFinancialPositionAbstract")}
  {_loc("l_tbl", "StatementTable")}
  {_loc("l_li", "StatementLineItems")}
  {_loc("l_cash", "CashAndCashEquivalentsAtCarryingValue")}
  {_loc("l_assets", "Assets")}
  {_arc("l_bs", "l_tbl", 1)}
  {_arc("l_tbl", "l_li", 1)}
  {_arc("l_li", "l_cash", 2)}
  {_arc("l_li", "l_assets", 1, _TOTAL)}
 </link:presentationLink>
</link:linkbase>"""

LAB_XML = f"""<link:linkbase {_XL}>
 <link:labelLink>
  {_loc("c_rev", "RevenueFromContractWithCustomerExcludingAssessedTax")}
  <link:labelArc xlink:from="c_rev" xlink:to="t_rev"/>
  <link:label xlink:label="t_rev" xlink:role="{_STD}">Net revenues</link:label>
  {_loc("c_cogs", "CostOfRevenue")}
  <link:labelArc xlink:from="c_cogs" xlink:to="t_cogs"/>
  <link:label xlink:label="t_cogs" xlink:role="{_STD}">Cost of revenue</link:label>
  <link:label xlink:label="t_cogs" xlink:role="{_NEGATED}">Cost of net revenues</link:label>
  {_loc("c_gp", "GrossProfit")}
  <link:labelArc xlink:from="c_gp" xlink:to="t_gp"/>
  <link:label xlink:label="t_gp" xlink:role="{_TOTAL}">Gross profit</link:label>
 </link:labelLink>
</link:linkbase>"""


def test_filing_summary_matches_three_statements_only():
    roles = parse_filing_summary(FILING_SUMMARY)
    assert set(roles) == {"income", "balance", "cashflow"}
    assert roles["income"] == (_ROLE_IS, "CONSOLIDATED STATEMENTS OF INCOME")
    assert roles["balance"][0] == _ROLE_BS      # Parenthetical excluded
    assert "Parenthetical" not in roles["balance"][1]


def test_statement_rows_order_depth_labels_totals():
    roles = {"income": (_ROLE_IS, "IS"), "balance": (_ROLE_BS, "BS")}
    rows = build_statement_rows(PRE_XML, LAB_XML, roles)

    inc = rows["income"]
    assert [r.concept for r in inc] == [
        "IncomeStatementAbstract",
        "RevenueFromContractWithCustomerExcludingAssessedTax",
        "CostOfRevenue", "GrossProfit", "NetIncomeLoss"]
    assert inc[0].is_abstract and inc[0].depth == 0
    assert inc[1].label == "Net revenues" and inc[1].depth == 1
    # preferredLabel resolves against the lab linkbase (negated variant)
    assert inc[2].label == "Cost of net revenues"
    assert inc[3].is_total and inc[3].label == "Gross profit"
    # no lab entry -> humanized concept name
    assert inc[4].label == "Net Income Loss" and inc[4].is_total

    bal = rows["balance"]
    # Table/LineItems scaffolding skipped, children promoted to depth 1;
    # the 'order' attribute (Assets=1, Cash=2) beats document order
    assert [r.concept for r in bal] == [
        "StatementOfFinancialPositionAbstract", "Assets",
        "CashAndCashEquivalentsAtCarryingValue"]
    assert bal[1].depth == 1 and bal[1].is_total


def test_annual_values_for_concept_units_and_lookup():
    facts = build_testco_companyfacts()
    annual = parse_companyfacts(facts, "TESTCO")
    fy_ends = annual.fy_ends
    vals, unit = annual_values_for_concept(
        facts, "RevenueFromContractWithCustomerExcludingAssessedTax", fy_ends)
    assert unit == "USD"
    assert vals[-1] == pytest.approx(REVENUE[2025])
    vals, unit = annual_values_for_concept(
        facts, "WeightedAverageNumberOfDilutedSharesOutstanding", fy_ends)
    assert unit == "shares"
    assert vals[-1] == pytest.approx(SHARES[2025])
    # instant concepts match balance-sheet dates
    vals, unit = annual_values_for_concept(facts, "Assets", fy_ends)
    assert vals[-1] == pytest.approx(REVENUE[2025] * 2.0)
    # unknown concept -> all None, no unit
    vals, unit = annual_values_for_concept(facts, "NoSuchConcept", fy_ends)
    assert vals == [None] * len(fy_ends) and unit == ""


def _data():
    facts = build_testco_companyfacts()
    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 5, 1))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(parse_companyfacts(facts, "TESTCO"), d)
    return d


def _stub_statements():
    return {
        "income": [
            PresRow("IncomeStatementAbstract", "Income Statement [Abstract]",
                    0, False, True),
            PresRow("RevenueFromContractWithCustomerExcludingAssessedTax",
                    "Net revenues", 1, False, False),
            PresRow("NetIncomeLoss", "Net income", 1, True, False),
            PresRow("WeightedAverageNumberOfDilutedSharesOutstanding",
                    "Diluted weighted average shares", 1, False, False),
        ],
        "balance": [
            PresRow("Assets", "Total assets", 1, True, False),
            PresRow("CashAndCashEquivalentsAtCarryingValue",
                    "Cash and cash equivalents", 1, False, False),
        ],
        "cashflow": [
            PresRow("NetCashProvidedByUsedInOperatingActivities",
                    "Net cash provided by operating activities",
                    1, True, False),
        ],
        "_short_names": {
            "income": "CONSOLIDATED STATEMENTS OF INCOME",
            "balance": "CONSOLIDATED BALANCE SHEETS",
            "cashflow": "CONSOLIDATED STATEMENTS OF CASH FLOWS"},
    }


def test_statement_sheets_written_as_filed(tmp_path):
    d = _data()
    d.statements = _stub_statements()
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    wb = load_workbook(str(out))
    assert wb.sheetnames[:4] == ["Financial Model", "Income Statement",
                                 "Balance Sheet", "Cash Flow"]

    ws = wb["Income Statement"]
    assert ws.freeze_panes == "B2"
    header = [c.value for c in ws[1]]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    fy25 = header.index("FY2025") + 1
    # abstract row: bold header, no values
    abs_row = labels.index("Income Statement [Abstract]") + 1
    assert ws.cell(row=abs_row, column=1).font.bold
    assert ws.cell(row=abs_row, column=fy25).value is None
    # depth-1 rows indent two spaces; money and share formats by unit
    rev_row = labels.index("  Net revenues") + 1
    assert ws.cell(row=rev_row, column=fy25).value == pytest.approx(
        REVENUE[2025] / 1e6)
    assert ws.cell(row=rev_row, column=fy25).number_format == _FMT_MONEY
    sh_row = labels.index("  Diluted weighted average shares") + 1
    assert ws.cell(row=sh_row, column=fy25).value == pytest.approx(
        SHARES[2025] / 1e6)
    assert ws.cell(row=sh_row, column=fy25).number_format == _FMT_SHARES
    ni_row = labels.index("  Net income") + 1
    assert ws.cell(row=ni_row, column=1).font.bold      # totalLabel row
    # KPI boundary sentence on the Income Statement sheet only
    joined = " ".join(str(v) for v in labels if v)
    assert "GMV, TPV, NIMAL" in joined
    bs_joined = " ".join(str(ws2.cell(row=r, column=1).value or "")
                         for ws2 in [wb["Balance Sheet"]]
                         for r in range(1, ws2.max_row + 1))
    assert "GMV" not in bs_joined
    assert "presentation linkbase" in bs_joined         # provenance footnote


def test_degrades_to_model_note_when_statements_missing(tmp_path):
    d = _data()
    d.statements = None
    d.statements_note = ("EdgarError: SEC Archives blocks the placeholder "
                         "User-Agent (HTTP 403). Set SEC_EDGAR_USER_AGENT "
                         "to 'name email' and retry.")
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    wb = load_workbook(str(out))
    assert wb.sheetnames == ["Financial Model"]         # no empty sheets
    labels = [wb.active.cell(row=r, column=1).value
              for r in range(1, wb.active.max_row + 1)]
    note = next(str(v) for v in labels
                if v and "As-filed statement sheets unavailable" in str(v))
    assert "SEC_EDGAR_USER_AGENT" in note


def test_segments_sheet_blocks_ties_and_footnote(tmp_path):
    from test_segments import _data_with_segments
    d = _data_with_segments()
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    wb = load_workbook(str(out))
    assert "Segments" in wb.sheetnames
    ws = wb["Segments"]
    assert ws.freeze_panes == "B2"
    header = [c.value for c in ws[1]]
    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert "Revenue by product / service" in labels
    assert "Revenue by geography" in labels
    fy25 = header.index("FY2025") + 1
    br = labels.index("  Brazil") + 1
    assert ws.cell(row=br, column=fy25).value == pytest.approx(1000.0)
    # one Σ/tie pair per Revenue axis (two Revenue blocks here)
    assert labels.count("   Σ members") == 2
    gap_rows = [i + 1 for i, v in enumerate(labels)
                if v == "   vs consolidated (gap %)"]
    assert len(gap_rows) == 2
    for gr in gap_rows:  # both axes are complete in the fixture -> 0.0%
        assert ws.cell(row=gr, column=fy25).value == pytest.approx(0.0)
    joined = " ".join(str(v) for v in labels if v)
    assert "fiscal spans as reported" in joined         # footnote present
