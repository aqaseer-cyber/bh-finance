"""FIX-11c: per-span quarterly gap-fill + LTM basis provenance."""
import datetime as dt

import pytest
from openpyxl import load_workbook

from conftest import CAPEX, CFO, build_testco_companyfacts
from forensic_viz.edgar import parse_companyfacts, parse_quarterly_facts
from forensic_viz.metrics import DashboardData, apply_track, build_fundamental_metrics
from forensic_viz.model_export import build_model_rows, export_financial_model
from test_model_export import _facts_with_quarters


def _dur(start, end, val, filed="2026-08-05"):
    return {"start": start, "end": end, "val": val, "form": "10-Q",
            "fp": "Q", "filed": filed}


def _with_secondary_capex(facts):
    """MELI shape: annual capex under …PropertyPlantAndEquipment, interim
    ONLY under …ProductiveAssets (whose annual history the winner beats)."""
    facts["facts"]["us-gaap"]["PaymentsToAcquireProductiveAssets"] = {
        "units": {"USD": [
            _dur("2026-01-01", "2026-03-31", 40e6, "2026-05-05"),
            _dur("2026-01-01", "2026-06-30", 85e6),
            _dur("2025-01-01", "2025-06-30", 70e6),
        ]}}
    return facts


def _data(facts):
    d = DashboardData(ticker="T", company="T Inc", subtitle="",
                      generated=dt.date(2026, 8, 10))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(parse_companyfacts(facts, "T"), d)
    return d


def test_secondary_tag_fills_quarters_with_source_note():
    facts = _with_secondary_capex(_facts_with_quarters())
    annual = parse_companyfacts(facts, "T")
    qdata = parse_quarterly_facts(facts, annual)
    assert any("capex: interim spans also from "
               "['PaymentsToAcquireProductiveAssets']" == n
               for n in qdata.source_notes)
    rows, _fy, _q = build_model_rows(annual, qdata)
    capex = rows["capex"]
    # Q1'26 discrete filed; Q2'26 derived by YTD differencing — both from
    # the secondary tag the old first-tag-with-data rule never reached
    assert capex.q[2:] == [pytest.approx(40e6), pytest.approx(45e6)]
    assert capex.ltm == pytest.approx(CAPEX[2025] + 85e6 - 70e6)
    assert capex.ltm_basis == "ltm"
    # both FCF legs now carry a true-LTM basis -> the derived LTM combines
    fcf = rows["=fcf"]
    assert fcf.ltm_basis == "ltm"
    assert fcf.ltm == pytest.approx(
        (CFO[2025] + 310e6 - 280e6) - (CAPEX[2025] + 85e6 - 70e6))


def test_mixed_basis_ltm_suppressed_with_footnote(tmp_path):
    # cfo has a true LTM (YTD + comparative); capex has NO current-year
    # interim under any tag -> its LTM falls back to FY -> =fcf is mixed
    d = _data(_facts_with_quarters())
    annual = d.fundamentals
    rows, _fy, _q = build_model_rows(
        annual, parse_quarterly_facts(annual.raw_facts, annual))
    assert rows["cfo"].ltm_basis == "ltm"
    assert rows["capex"].ltm_basis == "fy"
    assert rows["=fcf"].ltm is None and rows["=fcf"].ltm_basis == "mixed"

    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    ws = load_workbook(str(out))["Financial Model"]
    labels = [str(ws.cell(row=r, column=1).value or "")
              for r in range(1, ws.max_row + 1)]
    joined = " ".join(labels)
    assert "LTM suppressed (mixed basis): Free Cash Flow" in joined
    assert "LTM equals FY2025 (no current-year interim data" in joined
    assert "Capital Expenditure" in joined
    # instants tagged only at FY-ends are older than the newest quarter
    assert "older than the newest quarter column" in joined
    assert "Retained Earnings (2025-12-31)" in joined
    # gap-fill audit rides verbatim
    assert any(lbl.startswith("Interim gap-fill: ") for lbl in labels)


def test_both_fy_legs_combine():
    """No current-year interim at all: both legs are FY-basis — the
    derived FCF LTM is the fiscal year, not suppressed."""
    facts = build_testco_companyfacts()
    annual = parse_companyfacts(facts, "T")
    rows, _fy, q_ends = build_model_rows(
        annual, parse_quarterly_facts(facts, annual))
    assert q_ends[-1] == dt.date(2025, 12, 31)  # spine ends at the FY end
    assert rows["cfo"].ltm_basis == "fy"
    assert rows["capex"].ltm_basis == "fy"
    fcf = rows["=fcf"]
    assert fcf.ltm == pytest.approx(CFO[2025] - CAPEX[2025])
    assert fcf.ltm_basis == "fy"
