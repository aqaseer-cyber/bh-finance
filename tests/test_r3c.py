"""v3 R3c — THE workbook: expanded Cover (P1 mirror + run identity),
sheet order, the Audit sheet, and the principle-2 equality exercise
(Cover FV == report P1 == verdict object). All offline.
"""
import datetime as dt
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from conftest import build_testco_companyfacts
from forensic_viz.edgar import parse_companyfacts
from forensic_viz.metrics import (
    DashboardData, apply_track, build_fundamental_metrics,
    build_price_metrics,
)
from forensic_viz.model_export import export_financial_model
from forensic_viz.prices import PriceSeries
from forensic_viz.reconcile import AuditEntry, AuditReport
from forensic_viz.runid import run_identity
from forensic_viz.valuation import CaseInputs, ValuationInputs, build_valuation
from forensic_viz.verdict import build_verdict

FIXTURES = Path(__file__).parent / "fixtures"


def _testco(prices=True):
    d = DashboardData(ticker="TESTCO", company="TESTCO INC", subtitle="fx",
                      generated=dt.date(2026, 7, 19))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(
        parse_companyfacts(build_testco_companyfacts(), "TESTCO"), d)
    if prices:
        raw = json.loads((FIXTURES / "aapl_weekly_5y.json").read_text())
        build_price_metrics(PriceSeries(
            symbol="TESTCO",
            dates=[dt.date.fromisoformat(s) for s in raw["dates"]],
            closes=raw["close"], source="fixture"), d)
    return d


def _valued(d):
    inputs = ValuationInputs(
        method="dcf", discount_rate=0.09,
        cases={"Bear": CaseInputs(g0=0.02, g_term=0.02),
               "Base": CaseInputs(g0=0.05, g_term=0.025),
               "Bull": CaseInputs(g0=0.08, g_term=0.03)})
    res = build_valuation(d, inputs)
    return res, build_verdict(d, inputs, res, rating="Buy")


def _cover_rows(wb):
    cov = wb["Cover"]
    return {str(cov.cell(row=r, column=1).value or ""):
            cov.cell(row=r, column=2).value
            for r in range(1, cov.max_row + 1)}


def _col_a(ws):
    return [str(ws.cell(row=r, column=1).value or "")
            for r in range(1, ws.max_row + 1)]


def test_sheet_order_and_audit_always_present(tmp_path):
    d = _testco()
    out = tmp_path / "wb.xlsx"
    export_financial_model(d, str(out))
    wb = load_workbook(str(out))
    # no statements / no segments on this fixture — both declared, and
    # the Audit sheet closes the book regardless
    assert wb.sheetnames == ["Cover", "Financial Model", "Audit"]


def test_cover_mirrors_p1_and_carries_run_identity(tmp_path):
    d = _testco()
    res, v = _valued(d)
    out = tmp_path / "wb.xlsx"
    export_financial_model(d, str(out), res=res, verdict=v)
    rows = _cover_rows(load_workbook(str(out)))
    # principle 2 exercised: Cover FV == the verdict object == what the
    # report's P1 prints (float round-trips through XML at ~1 ulp)
    assert rows["FV average"] == pytest.approx(v.fv_avg, abs=1e-9)
    assert rows["MoS at P₀"] == pytest.approx(v.mos, abs=1e-12)
    assert rows["Stressed MoS"] == pytest.approx(v.stressed_mos, abs=1e-12)
    assert rows["P₀"] == pytest.approx(d.last_close)
    assert rows["Rating"] == "Buy"
    assert str(rows["Coherence gate"]).startswith(v.coherence)
    from forensic_viz.dashboard import render_decision
    texts = " ".join(t.get_text() for ax in render_decision(d, res, v).axes
                     for t in ax.texts)
    assert f"${v.fv_avg:,.2f}" in texts     # P1 prints the same number
    rid, ihash = run_identity(d, res)
    assert str(rows["Run"]).startswith(f"{rid} · inputs {ihash}")
    assert "one run, three artifacts" in str(rows["Artifacts"])
    assert "unchallenged" in str(rows["Base quality"])


def test_cover_without_valuation_is_declared(tmp_path):
    d = _testco(prices=False)
    out = tmp_path / "wb.xlsx"
    export_financial_model(d, str(out))
    rows = _cover_rows(load_workbook(str(out)))
    assert "no valuation attached" in str(rows["Decision"])
    assert "FV average" not in rows
    # a6 line still present
    assert "Statement sheets" in rows


def test_cover_base_quality_challenged_is_red_keyed(tmp_path):
    d = _testco()
    d.sic_code = "6199"          # a1 financial signature
    out = tmp_path / "wb.xlsx"
    export_financial_model(d, str(out))
    wb = load_workbook(str(out))
    cov = wb["Cover"]
    row = next(r for r in range(1, cov.max_row + 1)
               if cov.cell(row=r, column=1).value == "Base quality")
    assert "conditional on the base" in str(cov.cell(row=row, column=2).value)
    # red-keyed: the value cell carries the negative color
    assert cov.cell(row=row, column=2).font.color.rgb.endswith("B3402A")


def test_audit_sheet_full_tables_and_no_ellipsis(tmp_path):
    d = _testco()
    rep = AuditReport(checked=9, matched=7, sources=["FMP"])
    rep.entries.append(AuditEntry("Revenue", "FY2023", 1700e6, 1650e6,
                                  "FMP", "restated"))
    rep.entries.append(AuditEntry("SBC", "FY2025", None, 50e6,
                                  "FMP", "rescuable"))
    d.audit_report = rep
    d.health_notes.append("a warning that must survive verbatim")
    res, v = _valued(d)
    out = tmp_path / "wb.xlsx"
    export_financial_model(d, str(out), res=res, verdict=v)
    wb = load_workbook(str(out))
    aud = wb["Audit"]
    cells = [str(aud.cell(row=r, column=c).value or "")
             for r in range(1, aud.max_row + 1) for c in range(1, 7)]
    joined = "|".join(cells)
    assert "RESTATED (EDGAR carries the recast" in joined
    assert "RESCUABLE (EDGAR empty)" in joined
    assert "a warning that must survive verbatim" in joined
    # the full tag map is a real table: every concept present
    f = d.fundamentals
    for concept in f.tags_used:
        assert concept in joined
    # principle 5 holds in the workbook too: no ellipses anywhere
    for sheet in wb.sheetnames:
        for row in wb[sheet].iter_rows():
            for cell in row:
                assert "…" not in str(cell.value or ""), (sheet, cell)


def test_model_sheet_kept_lean_with_pointers(tmp_path):
    d = _testco()
    out = tmp_path / "wb.xlsx"
    export_financial_model(d, str(out))
    fm = _col_a(load_workbook(str(out))["Financial Model"])
    assert not any(x.startswith("XBRL tags:") for x in fm)
    assert not any(x.startswith("DATA AUDIT") for x in fm)
    assert not any(x.startswith("SEGMENTS (as filed)") for x in fm)
    assert any("Audit sheet carries" in x for x in fm)
    # the FM-cell-specific provenance notes STAY on the sheet's face
    assert any("Not investment advice." in x for x in fm)
