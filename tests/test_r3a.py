"""v3 R3a — engine-adjacent computations (docs/V3_R3_EXPORT_DESIGN.md).

a1 base-quality gate · a2 regime-trimmed exit multiple · a3 stale-series
KPI guard · a4 restatement-aware reconciliation · a5 house-config
precedence · a6 statement-name hardening. All offline.
"""
import datetime as dt
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from conftest import (
    REVENUE, REVENUE_FY2023_ORIGINAL, build_testco_companyfacts,
)
from forensic_viz import config
from forensic_viz.anchors import (
    BaseQuality, assess_base_quality, trimmed_mean,
)
from forensic_viz.edgar import parse_companyfacts
from forensic_viz.kpi import stale_note
from forensic_viz.metrics import (
    DashboardData, apply_track, build_fundamental_metrics,
)

STATIC = Path(__file__).resolve().parent.parent / "webui" / "static"


def _testco_dashboard() -> DashboardData:
    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 7, 18))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(
        parse_companyfacts(build_testco_companyfacts(), "TESTCO"), d)
    return d


# ------------------------------------------------- a1: base-quality gate

def _bare(sic="3571", accruals=None, ni=None, cfo=None, assets=None,
          raw_facts=None):
    d = DashboardData(ticker="T", company="T", subtitle="",
                      generated=dt.date(2026, 7, 18))
    d.sic_code = sic
    d.accruals_ratio = accruals or []
    d.net_income = ni or []
    d.cfo = cfo or []
    d.total_assets = assets or []
    d.fundamentals = SimpleNamespace(raw_facts=raw_facts)
    return d


def _receivable_facts(loan_val):
    return {"facts": {"us-gaap": {
        "LoansAndLeasesReceivableNetReportedAmount": {
            "units": {"USD": [{"val": loan_val}]}},
    }}}


def test_a1_meli_like_receivable_book_trips():
    """The financing float: CFO 4× NI + a material credit book."""
    d = _bare(sic="7389",                      # not a finance SIC
              accruals=[0.02, -0.30, -0.28, -0.35],  # 3y median −30%
              ni=[1.0e9], cfo=[4.2e9],
              assets=[20e9],
              raw_facts=_receivable_facts(4.0e9))    # 20% of assets
    q = assess_base_quality(d)
    assert isinstance(q, BaseQuality)
    assert q.financial_signature is True
    assert q.cfo_ni_ratio == pytest.approx(4.2)
    assert q.accruals_median_3y == pytest.approx(-0.30)
    assert q.challenged is True
    assert "CFO runs 4.2× net income" in q.text
    assert "3y median accruals -30%" in q.text
    assert "conditional on the base" in q.text


def test_a1_pypl_like_clean_base_does_not_trip():
    d = _bare(sic="7389",
              accruals=[0.03, -0.04, 0.02, -0.01],
              ni=[4.0e9], cfo=[5.4e9],               # 1.35×
              assets=[80e9],
              raw_facts=_receivable_facts(6.0e9))    # 7.5% — immaterial
    q = assess_base_quality(d)
    assert q.financial_signature is False
    assert q.challenged is False
    assert q.text == ""


def test_a1_finance_sic_alone_trips():
    q = assess_base_quality(_bare(sic="6199", ni=[1e9], cfo=[1.1e9]))
    assert q.financial_signature and q.challenged
    assert "conditional on the base" in q.text


def test_a1_cfo_ni_ratio_alone_trips_and_negative_ni_is_na():
    q = assess_base_quality(_bare(ni=[1e9], cfo=[3.5e9]))
    assert q.challenged and q.cfo_ni_ratio == pytest.approx(3.5)
    # NI <= 0: the ratio is meaningless, never fabricated
    q2 = assess_base_quality(_bare(ni=[-1e9], cfo=[3.5e9]))
    assert q2.cfo_ni_ratio is None and not q2.challenged


def test_a1_accruals_median_uses_last_three_years():
    # early relic years must not enter the 3y window
    q = assess_base_quality(_bare(accruals=[0.90, 0.01, 0.02, 0.01]))
    assert q.accruals_median_3y == pytest.approx(0.01)
    assert not q.challenged


# ------------------------------------------ a2: regime-trimmed exit check

def test_a2_trimmed_mean_drops_top_and_bottom_quintile():
    # 10 obs, k=2: the crash years and the bubble years fall out
    vals = [2.0, 30.0, 11.0, 12.0, 13.0, 12.5, 11.5, 12.2, 28.0, 3.0]
    core = sorted(vals)[2:-2]
    assert trimmed_mean(vals) == pytest.approx(sum(core) / len(core))
    # < 5 usable obs: nothing dropped — the plain mean
    assert trimmed_mean([10.0, 12.0, 14.0, None]) == pytest.approx(12.0)
    assert trimmed_mean([None, None]) is None
    assert trimmed_mean([]) is None


def test_a2_exit_check_carries_trimmed_keys_additively():
    from forensic_viz.valuation import HORIZON, exit_multiple_check
    d = DashboardData(ticker="T", company="T", subtitle="",
                      generated=dt.date(2026, 7, 18))
    d.ebit_reported = [380e6, 400e6]
    # a three-year elevated regime the median alone would erase:
    # raw median 12.0; trim drops 10.0 and 60.0, mean(11, 12, 40) = 21.0
    d.ev_ebit_fy = [10.0, 11.0, 12.0, 40.0, 60.0]
    out = exit_multiple_check(d, base_g0=0.06, g_term=0.02, rate=0.10,
                              bridge=6e8, shares=100e6, price=40.0)
    # raw keys byte-for-byte as before (FIX-16e contract untouched)
    assert out["multiple"] == pytest.approx(12.0)
    ebit5 = 400e6
    for i in range(1, 6):
        ebit5 *= 1 + (0.06 + (0.02 - 0.06) * (i - 1) / (HORIZON - 1))
    eq5 = (12.0 * ebit5 - 6e8) / 100e6
    assert out["ebit5"] == pytest.approx(ebit5)
    assert out["eq5_ps"] == pytest.approx(eq5)
    assert out["fv_today"] == pytest.approx(eq5 / 1.1 ** 5)
    assert out["return_5y"] == pytest.approx((eq5 / 40.0) ** 0.2 - 1)
    # trimmed variant: same arithmetic on the interquintile mean
    assert out["multiple_trimmed"] == pytest.approx(21.0)
    eq5_t = (21.0 * ebit5 - 6e8) / 100e6
    assert out["fv_today_trimmed"] == pytest.approx(eq5_t / 1.1 ** 5)
    assert out["return_5y_trimmed"] == pytest.approx(
        (eq5_t / 40.0) ** 0.2 - 1)


def test_a2_exit_check_trimmed_below_five_obs_is_plain_mean():
    from forensic_viz.valuation import exit_multiple_check
    d = DashboardData(ticker="T", company="T", subtitle="",
                      generated=dt.date(2026, 7, 18))
    d.ebit_reported = [400e6]
    d.ev_ebit_fy = [10.0, 12.0, 14.0]   # symmetric: mean == median
    out = exit_multiple_check(d, 0.06, 0.02, 0.10, 6e8, 100e6, 40.0)
    assert out["multiple_trimmed"] == pytest.approx(out["multiple"])
    assert out["fv_today_trimmed"] == pytest.approx(out["fv_today"])


# --------------------------------------------- a3: stale-series KPI guard

def test_a3_stale_note_padded_series():
    labels = ["FY2021", "FY2022", "FY2023", "FY2024"]
    # dies two years early -> the relic value must NOT render
    assert stale_note([1.0, 2.0, None, None], labels) == \
        "n/a (series ends FY2022)"
    # current through the latest FY -> show the value
    assert stale_note([1.0, None, None, 4.0], labels) is None
    # nothing tagged at all -> ordinary n/a handling, no year to cite
    assert stale_note([None, None, None, None], labels) is None
    assert stale_note(None, labels) is None
    assert stale_note([1.0], []) is None


def test_a3_stale_note_tail_aligned_short_and_long_series():
    labels = ["FY2021", "FY2022", "FY2023", "FY2024"]
    # shorter series aligns at the tail (chart-join convention)
    assert stale_note([3.0, 4.0], labels) is None
    assert stale_note([3.0, None], labels) == "n/a (series ends FY2023)"
    # longer series whose last value predates the whole label window
    assert stale_note([1.0, None, None, None, None, None], labels) == \
        "n/a (series ends before FY2021)"


def test_a3_overview_js_mirrors_the_guard():
    js = (STATIC / "overview.js").read_text(encoding="utf-8")
    assert "staleNote" in js
    assert "series ends" in js
    # the FCF ex-SBC relic case: the yield tile guards on the series
    assert "fcf_ex_sbc" in js


# ------------------------------- a4: restatement-aware reconciliation

def test_a4_edgar_restated_detects_the_amended_span():
    from forensic_viz.reconcile import _edgar_restated
    f = parse_companyfacts(build_testco_companyfacts(), "TESTCO")
    # FY2023 was filed twice (10-K then 10-K/A, Δ$50M > tolerance)
    assert _edgar_restated(f, "revenue", dt.date(2023, 12, 31), 2e6)
    # FY2024 was filed once — no restatement
    assert not _edgar_restated(f, "revenue", dt.date(2024, 12, 31), 2e6)
    # unknown concept/tag — never crashes, never flags
    assert not _edgar_restated(f, "goodwill", dt.date(2024, 12, 31), 2e6)


def test_a4_reconcile_classifies_restated_separately():
    from forensic_viz.reconcile import AuditReport, reconcile_fmp
    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 7, 18))
    d.fundamentals = parse_companyfacts(build_testco_companyfacts(),
                                        "TESTCO")
    statements = {"income": [
        # provider still serves the ORIGINAL FY2023 number (the MELI
        # false-alarm shape) -> restated, not divergent
        {"date": "2023-12-31", "revenue": REVENUE_FY2023_ORIGINAL},
        # a genuinely wrong FY2024 number -> divergent as before
        {"date": "2024-12-31", "revenue": REVENUE[2024] * 1.5},
    ], "cashflow": [], "balance": []}
    rep = AuditReport()
    reconcile_fmp(d, statements, rep)
    kinds = {(e.item, e.fy): e.kind for e in rep.entries}
    assert kinds[("Revenue", "FY2023")] == "restated"
    assert kinds[("Revenue", "FY2024")] == "divergent"
    assert len(rep.restated) == 1 and len(rep.divergent) == 1
    assert "1 divergent" in rep.summary()
    assert "1 restated" in rep.summary()


def test_a4_export_prints_the_restated_status(tmp_path):
    from forensic_viz.model_export import export_financial_model
    from forensic_viz.reconcile import AuditEntry, AuditReport
    d = _testco_dashboard()
    rep = AuditReport(checked=5, matched=4, sources=["FMP"])
    rep.entries.append(AuditEntry("Revenue", "FY2023", 1700e6, 1650e6,
                                  "FMP", "restated"))
    d.audit_report = rep
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    ws = load_workbook(str(out))["Financial Model"]
    cells = [str(ws.cell(row=r, column=c).value or "")
             for r in range(1, ws.max_row + 1) for c in range(1, 7)]
    assert any("RESTATED (EDGAR carries the recast; provider carries "
               "the original)" in v for v in cells)


# ----------------------------------------- a5: house-config precedence

def test_a5_house_file_reaches_erp_after_settings_apply(tmp_path,
                                                        monkeypatch):
    """The field bug: apply_user_settings updated USER_HOUSE_FILE but the
    ASSUMPTION constants kept their import-time values."""
    monkeypatch.delenv("HOUSE_ASSUMPTIONS_FILE", raising=False)
    monkeypatch.chdir(tmp_path)  # no stray house_assumptions.toml
    p = tmp_path / "house.toml"
    p.write_text("erp = 0.055\ngdp_cap = 0.030\n", encoding="utf-8")
    try:
        config.apply_user_settings({"house_file": str(p)})
        assert config.HOUSE_LOADED is True
        assert config.ERP_ASSUMPTION == pytest.approx(0.055)
        assert config.GDP_CAP == pytest.approx(0.030)
        # clearing the setting restores the code defaults
        config.apply_user_settings({})
        assert config.HOUSE_LOADED is False
        assert config.ERP_ASSUMPTION == pytest.approx(0.046)
        assert config.GDP_CAP == pytest.approx(0.035)
    finally:
        config.apply_user_settings({})


def test_a5_waccbuild_default_reads_config_at_build_time(tmp_path,
                                                         monkeypatch):
    """The deeper cause: WaccBuild.erp was a plain dataclass default,
    frozen at import — a later house apply never reached a new build."""
    from forensic_viz.rates import WaccBuild
    monkeypatch.delenv("HOUSE_ASSUMPTIONS_FILE", raising=False)
    monkeypatch.chdir(tmp_path)
    p = tmp_path / "house.toml"
    p.write_text("erp = 0.061\n", encoding="utf-8")
    try:
        config.apply_user_settings({"house_file": str(p)})
        assert WaccBuild().erp == pytest.approx(0.061)
        config.apply_user_settings({})
        assert WaccBuild().erp == pytest.approx(0.046)
    finally:
        config.apply_user_settings({})


def test_a5_workbook_b5_equals_config_erp_at_fill_time(tmp_path,
                                                       monkeypatch):
    from forensic_viz.rates import WaccBuild
    from forensic_viz.workbook import fill_workbook
    monkeypatch.delenv("HOUSE_ASSUMPTIONS_FILE", raising=False)
    p = tmp_path / "house.toml"
    p.write_text("erp = 0.052\n", encoding="utf-8")
    d = _testco_dashboard()
    try:
        monkeypatch.chdir(tmp_path)
        config.apply_user_settings({"house_file": str(p)})
        d.wacc_build = WaccBuild(r_f=0.04)
        out = tmp_path / "wb.xlsx"
        fill_workbook(d, str(out))
        wb = load_workbook(str(out))
        assert wb["WACC_Build"]["B5"].value == pytest.approx(0.052)
        assert wb["WACC_Build"]["B5"].value == pytest.approx(
            config.ERP_ASSUMPTION)
    finally:
        config.apply_user_settings({})


# ------------------------------------- a6: statement-name hardening

_SUMMARY_TMPL = """<FilingSummary><MyReports>{reports}
</MyReports></FilingSummary>"""


def _report(short, role):
    return (f"<Report><ShortName>{short}</ShortName>"
            f"<Role>{role}</Role></Report>")


def test_a6_condensed_and_loss_variants_match():
    from forensic_viz.edgar import parse_filing_summary
    xml = _SUMMARY_TMPL.format(reports=(
        _report("CONDENSED CONSOLIDATED BALANCE SHEETS", "r/bs")
        + _report("CONSOLIDATED STATEMENTS OF OPERATIONS (LOSS)", "r/is")
        + _report("CONDENSED CONSOLIDATED STATEMENTS OF CASH FLOW",
                  "r/cf")))
    roles = parse_filing_summary(xml)
    assert set(roles) == {"income", "balance", "cashflow"}
    assert roles["income"][0] == "r/is"


def test_a6_comprehensive_only_fallback_for_income():
    from forensic_viz.edgar import parse_filing_summary
    # combined-statement filer: no separate income statement report
    xml = _SUMMARY_TMPL.format(reports=(
        _report("CONSOLIDATED BALANCE SHEETS", "r/bs")
        + _report("CONSOLIDATED STATEMENTS OF COMPREHENSIVE INCOME",
                  "r/ci")))
    roles = parse_filing_summary(xml)
    assert roles["income"][0] == "r/ci"
    # ordinary filer: the separate income statement still wins and the
    # comprehensive variant stays excluded (no duplicate)
    xml2 = _SUMMARY_TMPL.format(reports=(
        _report("CONSOLIDATED STATEMENTS OF INCOME", "r/is")
        + _report("CONSOLIDATED STATEMENTS OF COMPREHENSIVE INCOME",
                  "r/ci")))
    assert parse_filing_summary(xml2)["income"][0] == "r/is"


def test_a6_pre_linkbase_role_fallback():
    from forensic_viz.edgar import roles_from_pre_linkbase
    xl = ('xmlns:link="http://www.xbrl.org/2003/linkbase" '
          'xmlns:xlink="http://www.w3.org/1999/xlink"')
    pre = f"""<link:linkbase {xl}>
     <link:presentationLink
       xlink:role="http://x.example/role/StatementsOfIncome"/>
     <link:presentationLink
       xlink:role="http://x.example/role/ConsolidatedBalanceSheets"/>
     <link:presentationLink
       xlink:role="http://x.example/role/BalanceSheetsParenthetical"/>
     <link:presentationLink
       xlink:role="http://x.example/role/StatementsOfCashFlows"/>
    </link:linkbase>"""
    roles = roles_from_pre_linkbase(pre)
    assert set(roles) == {"income", "balance", "cashflow"}
    assert roles["income"][0].endswith("StatementsOfIncome")
    assert "(pre-linkbase fallback)" in roles["income"][1]
    assert "Parenthetical" not in roles["balance"][0]
    assert roles_from_pre_linkbase("<not xml") == {}


def test_a6_cover_declares_statement_completeness(tmp_path):
    from forensic_viz.model_export import (
        _statement_completeness_line, export_financial_model,
    )
    d = _testco_dashboard()
    # no statement structure fetched at all
    assert "not fetched" in _statement_completeness_line(d)
    # partial: the declared-missing note passes through verbatim
    d.statements = {"balance": [object()], "cashflow": [object()]}
    d.statements_note = ("Income Statement: not identified in this "
                         "filing's presentation — sheet omitted")
    assert _statement_completeness_line(d) == d.statements_note
    # all three present
    d.statements = {"income": [1], "balance": [1], "cashflow": [1]}
    assert "all identified" in _statement_completeness_line(d)
    # and the Cover carries the row
    d.statements = None
    d.statements_note = ""
    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    cover = load_workbook(str(out))["Cover"]
    rows = {str(cover.cell(row=r, column=1).value or ""):
            str(cover.cell(row=r, column=2).value or "")
            for r in range(1, cover.max_row + 1)}
    assert "Statement sheets" in rows
    assert "not fetched" in rows["Statement sheets"]


def test_a6_declared_missing_note_from_fetch(monkeypatch):
    """fetch_statement_presentation declares what it could not identify
    instead of failing the whole enrichment."""
    import forensic_viz.edgar as edgar

    summary_xml = _SUMMARY_TMPL.format(reports=_report(
        "CONSOLIDATED BALANCE SHEETS", "http://x.example/role/BS"))
    xl = ('xmlns:link="http://www.xbrl.org/2003/linkbase" '
          'xmlns:xlink="http://www.w3.org/1999/xlink"')
    pre_xml = f"""<link:linkbase {xl}>
     <link:presentationLink xlink:role="http://x.example/role/BS">
      <link:loc xlink:label="l_a" xlink:href="x.xsd#us-gaap_Assets"/>
      <link:loc xlink:label="l_r"
        xlink:href="x.xsd#us-gaap_StatementOfFinancialPositionAbstract"/>
      <link:presentationArc xlink:from="l_r" xlink:to="l_a" order="1"/>
     </link:presentationLink>
    </link:linkbase>"""

    class FakeSession:
        def __init__(self, cache):
            pass

        def get_text(self, url, ttl):
            if url.endswith("FilingSummary.xml"):
                return summary_xml
            if url.endswith("_pre.xml"):
                return pre_xml
            raise RuntimeError("no lab linkbase")

    monkeypatch.setattr(edgar, "_SecSession", FakeSession)
    monkeypatch.setattr(edgar, "_require_declared_ua", lambda: None)
    annual = SimpleNamespace(cik=1234567,
                             latest_10k_accession="0000000000-26-000001",
                             latest_10k_document="x-20251231.htm")
    rows, notes = edgar.fetch_statement_presentation(annual, cache=None)
    assert "balance" in rows and "income" not in rows
    assert any(n == "Income Statement: not identified in this filing's "
                    "presentation — sheet omitted" for n in notes)
    assert any(n.startswith("Cash Flow: not identified") for n in notes)
