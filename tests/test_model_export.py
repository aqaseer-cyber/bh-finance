"""Financial-model export: last-4-quarter consolidation + the one-sheet
workbook with % change rows.

Extends the TESTCO fixture with FY2026 interim filings on top of the
existing FY2025 quarterly noise (Q1–Q3'25 discrete revenue):
- revenue: discrete 3-month values for Q1'26/Q2'26, the H1'26 fiscal-YTD
  span, the 9-month FY2025 YTD (for the derived Q4'25 = FY − 9M), and the
  year-ago H1'25 comparative (for LTM);
- cfo: YTD-ONLY spans (like a real 10-Q cash-flow statement), so Q2'26
  must be derived by differencing;
- instants (cash, total assets): balances at both 2026 quarter ends.
"""
import datetime as dt

import pytest
from openpyxl import load_workbook

from conftest import CFO, REVENUE, build_testco_companyfacts
from forensic_viz.edgar import parse_companyfacts, parse_quarterly_facts
from forensic_viz.metrics import DashboardData, apply_track, build_fundamental_metrics
from forensic_viz.model_export import (
    build_model_rows, export_financial_model, quarter_label,
)

# FY2026 interim values (fixture FYs end 12-31; FY2025 quarters are 475 each)
Q_2025 = REVENUE[2025] / 4                      # 475e6
Q1_REV, Q2_REV = 520e6, 540e6
Q1_CFO, H1_CFO = 150e6, 310e6
H1_REV_PRIOR = 2 * Q_2025                       # consistent with the noise
H1_CFO_PRIOR = 280e6
Q1_CASH, Q2_CASH = 500e6, 520e6


def _dur(start, end, val, filed):
    return {"start": start, "end": end, "val": val, "form": "10-Q",
            "fp": "Q", "filed": filed}


def _inst(end, val, filed):
    return {"end": end, "val": val, "form": "10-Q", "fp": "Q", "filed": filed}


def _facts_with_quarters() -> dict:
    facts = build_testco_companyfacts()
    gaap = facts["facts"]["us-gaap"]
    rev = gaap["RevenueFromContractWithCustomerExcludingAssessedTax"]["units"]["USD"]
    rev += [
        _dur("2025-01-01", "2025-09-30", 3 * Q_2025, "2025-11-05"),  # 9M YTD
        _dur("2026-01-01", "2026-03-31", Q1_REV, "2026-05-05"),      # Q1 3M
        _dur("2026-04-01", "2026-06-30", Q2_REV, "2026-08-05"),      # Q2 3M
        _dur("2026-01-01", "2026-06-30", Q1_REV + Q2_REV, "2026-08-05"),
        _dur("2025-01-01", "2025-06-30", H1_REV_PRIOR, "2026-08-05"),
    ]
    cfo = gaap["NetCashProvidedByUsedInOperatingActivities"]["units"]["USD"]
    cfo += [  # YTD-only, like a real 10-Q cash-flow statement
        _dur("2026-01-01", "2026-03-31", Q1_CFO, "2026-05-05"),
        _dur("2026-01-01", "2026-06-30", H1_CFO, "2026-08-05"),
        _dur("2025-01-01", "2025-06-30", H1_CFO_PRIOR, "2026-08-05"),
    ]
    gaap["CashAndCashEquivalentsAtCarryingValue"]["units"]["USD"] += [
        _inst("2026-03-31", Q1_CASH, "2026-05-05"),
        _inst("2026-06-30", Q2_CASH, "2026-08-05"),
    ]
    gaap["Assets"]["units"]["USD"] += [
        _inst("2026-03-31", 4.0e9, "2026-05-05"),
        _inst("2026-06-30", 4.1e9, "2026-08-05"),
    ]
    # a per-share row so the FIX-12h format map is exercised end to end
    gaap["EarningsPerShareDiluted"] = {"units": {"USD/shares": [
        {"start": f"{y}-01-01", "end": f"{y}-12-31", "val": 3.0 + 0.5 * i,
         "fy": y + 1, "fp": "FY", "form": "10-K",
         "filed": f"{y + 1}-02-15", "frame": f"CY{y}"}
        for i, y in enumerate(range(2024, 2026))
    ]}}
    return facts


def _data(facts) -> DashboardData:
    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 8, 10))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(parse_companyfacts(facts, "TESTCO"), d)
    return d


def test_last_four_quarters_span_the_fy_boundary():
    facts = _facts_with_quarters()
    annual = parse_companyfacts(facts, "TESTCO")
    qdata = parse_quarterly_facts(facts, annual)
    rows, fy_ends, q_ends = build_model_rows(annual, qdata)

    assert q_ends == [dt.date(2025, 9, 30), dt.date(2025, 12, 31),
                      dt.date(2026, 3, 31), dt.date(2026, 6, 30)]
    assert [quarter_label(q, fy_ends) for q in q_ends] == \
        ["Q3'25", "Q4'25", "Q1'26", "Q2'26"]

    rev = rows["revenue"]
    # Q3'25 filed discrete; Q4'25 derived = FY2025 − 9M YTD; Q1/Q2'26 filed
    assert rev.q == [pytest.approx(Q_2025), pytest.approx(Q_2025),
                     pytest.approx(Q1_REV), pytest.approx(Q2_REV)]
    assert rev.ltm == pytest.approx(
        REVENUE[2025] + (Q1_REV + Q2_REV) - H1_REV_PRIOR)
    # year-ago quarters (drive the YoY cells): Q3'24/Q4'24 underivable in
    # the fixture; Q1'25/Q2'25 are the filed 475s
    assert rev.q_prior == [None, None, pytest.approx(Q_2025),
                           pytest.approx(Q_2025)]

    # cfo: no discrete spans — Q2'26 derived by YTD differencing; the two
    # 2025 quarters have no CFO interim data at all
    cfo = rows["cfo"]
    assert cfo.q == [None, None, pytest.approx(Q1_CFO),
                     pytest.approx(H1_CFO - Q1_CFO)]
    assert cfo.ltm == pytest.approx(CFO[2025] + H1_CFO - H1_CFO_PRIOR)

    # balance sheet: Q4'25 column carries the FY-end balance from the 10-K
    cash = rows["cash"]
    assert cash.q[1] == pytest.approx(REVENUE[2025] * 0.25)
    assert cash.q[2:] == [pytest.approx(Q1_CASH), pytest.approx(Q2_CASH)]
    assert cash.ltm == pytest.approx(Q2_CASH)


def test_derived_q4_and_ltm_when_latest_period_is_fy_end():
    """Right after a 10-K: quarters = Q1..Q4 of the completed FY."""
    facts = build_testco_companyfacts()  # noise = Q1–Q3'25 discrete revenue
    annual = parse_companyfacts(facts, "TESTCO")
    rows, fy_ends, q_ends = build_model_rows(
        annual, parse_quarterly_facts(facts, annual))
    assert [quarter_label(q, fy_ends) for q in q_ends] == \
        ["Q1'25", "Q2'25", "Q3'25", "Q4'25"]
    # Q4 falls back to FY − ΣQ1..Q3 when no 9M YTD span is filed
    assert rows["revenue"].q == [pytest.approx(Q_2025)] * 4
    # latest period end IS the FY end -> LTM is simply the fiscal year
    assert rows["revenue"].ltm == pytest.approx(REVENUE[2025])


def test_export_layout_adapts_and_carries_pct_rows(tmp_path):
    facts = _facts_with_quarters()
    d = _data(facts)
    out = tmp_path / "model.xlsx"
    export_financial_model(d, str(out))

    ws = load_workbook(str(out))["Financial Model"]
    header = [c.value for c in ws[1]]
    assert header[0] == "Line Items" and header[-1] == "LTM"
    assert header[-5:-1] == ["Q3'25", "Q4'25", "Q1'26", "Q2'26"]
    assert ws.freeze_panes == "B2"

    labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    # FIX-12h: units live in the section headers
    for section in ("INCOME STATEMENT ($mm; EPS in $, shares in mm)",
                    "BALANCE SHEET (period end, $mm)",
                    "CASH FLOW STATEMENT ($mm)"):
        assert section in labels
    # adaptive layout: untagged lines are dropped, tagged ones kept
    assert "Selling, General & Administrative" not in labels
    assert "Sales & Marketing" not in labels
    assert "Inventory" in labels

    ltm_col = header.index("LTM") + 1
    rev_row = labels.index("Total Revenue") + 1
    assert ws.cell(row=rev_row, column=ltm_col).value == pytest.approx(
        (REVENUE[2025] + Q1_REV + Q2_REV - H1_REV_PRIOR) / 1e6)

    # % change row directly under Total Revenue: FY cells YoY, quarter
    # cells YoY vs the same fiscal quarter a year earlier, LTM cell blank
    assert labels[rev_row] == "   % change"
    pct_row = rev_row + 1
    fy25_col = header.index("FY2025") + 1
    assert ws.cell(row=pct_row, column=fy25_col).value == pytest.approx(
        REVENUE[2025] / REVENUE[2024] - 1)            # annual YoY
    assert ws.cell(row=pct_row, column=header.index("Q1'26") + 1).value == \
        pytest.approx(Q1_REV / Q_2025 - 1)            # YoY vs Q1'25
    assert ws.cell(row=pct_row, column=header.index("Q2'26") + 1).value == \
        pytest.approx(Q2_REV / Q_2025 - 1)            # YoY vs Q2'25
    # Q3'24 is underivable in the fixture -> the Q3'25 YoY cell stays blank
    assert ws.cell(row=pct_row, column=header.index("Q3'25") + 1).value is None
    assert ws.cell(row=pct_row, column=ltm_col).value is None

    # FIX-12h: number formats by row kind (money / per-share / shares / %)
    from forensic_viz.model_export import (
        _FMT_MONEY, _FMT_PCT, _FMT_PS, _FMT_SHARES,
    )
    def fmt_of(label):
        row = labels.index(label) + 1
        return ws.cell(row=row, column=fy25_col).number_format
    assert fmt_of("Total Revenue") == _FMT_MONEY
    assert fmt_of("Diluted EPS ($)") == _FMT_PS
    assert fmt_of("Diluted Shares (mm)") == _FMT_SHARES
    assert ws.cell(row=pct_row, column=fy25_col).number_format == _FMT_PCT


def test_export_without_fundamentals_raises(tmp_path):
    d = DashboardData(ticker="X", company="X", subtitle="",
                      generated=dt.date(2026, 8, 10))
    with pytest.raises(ValueError):
        export_financial_model(d, str(tmp_path / "x.xlsx"))
