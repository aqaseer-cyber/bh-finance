"""FIX-16a: market-joined per-FY ratio series — hand-computed fixtures.

Every ratio must mask (None) on a missing or non-positive denominator;
market cap uses the same FY-end-close × diluted-shares MVE as Altman."""
import datetime as dt

import pytest

from forensic_viz.market import compute_market_ratios, summary_stat
from forensic_viz.metrics import DashboardData


def _d():
    d = DashboardData(ticker="T", company="T Inc", subtitle="",
                      generated=dt.date(2026, 7, 12))
    d.fy_ends = [dt.date(2023, 12, 31), dt.date(2024, 12, 31),
                 dt.date(2025, 12, 31)]
    d.fy_labels = ["FY2023", "FY2024", "FY2025"]
    # daily closes land exactly on the FY ends
    d.price_dates = list(d.fy_ends)
    d.price_closes = [50.0, 60.0, 80.0]
    d.last_close = 80.0
    d.diluted_shares = [100e6, 100e6, 100e6]
    d.eps_diluted = [2.0, -1.0, 4.0]          # loss year masks P/E
    d.ebit_reported = [400e6, 500e6, 640e6]
    d.total_debt = [1e9, 1e9, 1e9]
    d.cash = [4e8, 6e8, 2e8]
    d.minority_interest = [None, None, 50e6]
    d.preferred_equity = [None, None, None]
    d.fcf_ex_sbc = [300e6, 350e6, 400e6]
    d.book_equity = [2e9, 2.2e9, 2.5e9]
    d.goodwill = [3e8, 3e8, 3e8]
    d.intangibles = [None, 1e8, 1e8]
    d.dividends_paid = [None, 50e6, 60e6]
    d.buybacks = [None, None, 200e6]
    return d


def test_market_joined_series_hand_computed():
    d = _d()
    compute_market_ratios(d)
    assert d.market_cap_fy == [pytest.approx(5e9), pytest.approx(6e9),
                               pytest.approx(8e9)]
    assert d.net_debt_fy == [pytest.approx(6e8), pytest.approx(4e8),
                             pytest.approx(8e8)]
    # EV adds the bridge legs (MI in FY2025)
    assert d.ev_fy == [pytest.approx(5.6e9), pytest.approx(6.4e9),
                       pytest.approx(8.85e9)]
    assert d.pe_fy[0] == pytest.approx(25.0)
    assert d.pe_fy[1] is None                       # negative EPS masked
    assert d.pe_fy[2] == pytest.approx(20.0)
    assert d.ev_ebit_fy[0] == pytest.approx(5.6e9 / 400e6)
    assert d.net_debt_ebit_fy[2] == pytest.approx(8e8 / 640e6)
    assert d.adj_fcf_yield_fy[2] == pytest.approx(400e6 / 8e9)
    # tangible book: missing intangibles leg counts as 0
    assert d.tangible_book[0] == pytest.approx(2e9 - 3e8)
    assert d.tangible_book[1] == pytest.approx(2.2e9 - 3e8 - 1e8)
    # KPIs on the current market cap
    assert d.adj_fcf_yield_now == pytest.approx(400e6 / 8e9)
    assert d.owners_yield == pytest.approx((60e6 + 200e6) / 8e9)


def test_market_series_mask_without_prices_or_shares():
    d = _d()
    d.price_dates, d.price_closes = [], []
    d.fy_prices = []
    compute_market_ratios(d)
    assert d.market_cap_fy == [None, None, None]
    assert d.pe_fy == [None, None, None]
    assert d.ev_fy == [None, None, None]
    # net debt and tangible book need no price — they survive
    assert d.net_debt_fy[0] == pytest.approx(6e8)
    assert d.tangible_book[2] == pytest.approx(2.5e9 - 3e8 - 1e8)
    assert d.adj_fcf_yield_now is None or d.last_close  # last_close path
    d2 = _d()
    d2.last_close = None
    compute_market_ratios(d2)
    assert d2.owners_yield is None and d2.adj_fcf_yield_now is None


def test_summary_stat_cagr_and_avg():
    assert summary_stat([100.0, None, 121.0], "cagr") is None  # < 3 points
    assert summary_stat([100.0, 110.0, 121.0], "cagr") == pytest.approx(0.10)
    assert summary_stat([None, 100.0, 110.0, 121.0], "cagr") == \
        pytest.approx(0.10)                       # None-skipping, span-true
    assert summary_stat([-5.0, 100.0, 121.0], "cagr") is None  # endpoint ≤ 0
    assert summary_stat([0.10, 0.20, 0.30], "avg") == pytest.approx(0.20)
    assert summary_stat([0.10, None, 0.30], "avg") is None     # < 3 values


def test_export_market_block_and_summary_column(tmp_path):
    """FIX-16b: the export grows a MARKET & RATIOS block (FY columns +
    today in the LTM column) and a per-row CAGR/avg summary column."""
    import datetime as dt

    from openpyxl import load_workbook

    from forensic_viz.edgar import parse_companyfacts
    from forensic_viz.metrics import (
        apply_track, attach_fy_prices, build_fundamental_metrics,
    )
    from forensic_viz.model_export import export_financial_model
    from test_model_export import _facts_with_quarters

    d = DashboardData(ticker="TESTCO", company="TESTCO Inc", subtitle="",
                      generated=dt.date(2026, 8, 10))
    d.sic_code = "3571"
    apply_track(d, "auto")
    build_fundamental_metrics(
        parse_companyfacts(_facts_with_quarters(), "TESTCO"), d)
    d.price_dates = [e for e in d.fy_ends]
    d.price_closes = [50.0 + i for i in range(len(d.fy_ends))]
    d.last_close = d.price_closes[-1]
    attach_fy_prices(d)
    compute_market_ratios(d)
    assert any(v is not None for v in d.market_cap_fy)

    out = tmp_path / "m.xlsx"
    export_financial_model(d, str(out))
    ws = load_workbook(str(out))["Financial Model"]
    header = [c.value for c in ws[1]]
    assert header[-1] == "CAGR/avg"
    labels = [str(ws.cell(row=r, column=1).value or "")
              for r in range(1, ws.max_row + 1)]
    mk = next(i + 1 for i, l in enumerate(labels)
              if l.startswith("MARKET & RATIOS"))
    sum_col = len(header)
    fy_last_col = 1 + len(d.fundamentals.fy_ends)  # last FY column (1-based)

    def row_of(label):
        return next(i + 1 for i, l in enumerate(labels) if l == label)

    mcap_row = row_of("Market Cap")
    expect_mcap = d.market_cap_fy[-1] / 1e6
    assert ws.cell(row=mcap_row, column=fy_last_col).value == \
        pytest.approx(expect_mcap)
    # today's value rides in the LTM column
    ltm_col = header.index("LTM") + 1
    shares_now = next(v for v in reversed(d.diluted_shares) if v)
    assert ws.cell(row=mcap_row, column=ltm_col).value == \
        pytest.approx(d.last_close * shares_now / 1e6)
    # revenue row carries a CAGR summary matching summary_stat
    rev_row = next(i + 1 for i, l in enumerate(labels)
                   if l in ("Total Revenue",))
    rev_ann = [ws.cell(row=rev_row, column=c).value
               for c in range(2, fy_last_col + 1)]
    got = ws.cell(row=rev_row, column=sum_col).value
    assert got == pytest.approx(summary_stat(rev_ann, "cagr"))
    # footnote provenance
    joined = " ".join(labels)
    assert "MARKET & RATIOS: market cap = FY-end close" in joined
